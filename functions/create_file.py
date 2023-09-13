import os
import sys
from collections import namedtuple

from openpyxl.utils.cell import get_column_letter
import pandas as pd
import win32com.client
import win32timezone

from settings.templates import SALES_METRES_WITH_FIRST_COLUMN, SALES_METRES_WITH_OTHER_COLUMN, CORRECTED_METRES, \
    SALES_METRES_WITHOUT_FIRST_COLUMN, SALES_METRES_WITHOUT_OTHER_COLUMN, SALES_MONEY_WITH_FIRST_COLUMN, \
    CORRECTED_SALES, SALES_MONEY_WITHOUT_FIRST_COLUMN, SALES_MONEY_WITHOUT_OTHER_COLUMN, SALES_MONEY_WITH_OTHER_COLUMN, \
    PARTNER_SALES_METRES_FIRST_COLUMN, PARTNER_SALES_METRES_OTHER_COLUMN, PARTNER_SALES_MONEY_OTHER_COLUMN, \
    PARTNER_SALES_MONEY_FIRST_COLUMN, FORMULA_METRES_CLOSE, FORMULA_METRES_DENIAL, FORMULA_MONEY_CLOSE, \
    FORMULA_MONEY_DENIAL

period = '1_2014'
PRJ = 'АЛХИМОВО'
df = pd.DataFrame({'Проект': ['АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО' , 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО' , 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО' , 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО'], 'Очередь': [1, 1, 2, 2, 2, 2, 3, 3, 4, 4, 4, 4, 4, 4, 5, 5],
                   'Дом': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]})


ADD_COLUMNS = ['Полугодие','Год','Квартал_Год','Очередь','Дом','Тип','Контрагент','Тип контрагента','Договор','Сумма']
BIT_COLUMNS = ['Дата', 'Операция', 'Операция_1', 'Аналитика_Дт', 'Аналитика_Дт', 'Аналитика_Дт', 'Аналитика_Кт',
               'Аналитика_Кт', 'Аналитика_Кт', 'Аналитика_Кт', 'Счет_Дт', 'Сумма_Дт', 'Счет_Кт', 'Сумма_Кт', 'Тип операции', 'Сальдо накопительным итогом']
ADD_CRM_COLUMNS = ['Квартал регистрации договора', 'Год регистрации', 'Квартал_Год регистрации',
                            'Квартал расторжения договора', 'Год расторжения', 'Квартал_Год расторжения',
                            'Тип договора',
                            'Очередь', 'Дом', 'Учитывается(нет/да)', 'Контрагент', 'Тип контрагента (Партнер или нет)',
                            'Договор', 'Площадь', 'Сумма', 'Комментарий', 'Расторгнут?(да/нет)', 'Корректировка м.2',
                            'Корректировка тыс.руб.']

CRM_COLUMNS = ['Договор, #', '№ Договора', 'Тип договора', 'Дата заключения', 'Дата_рег договора',
                      'Дата раст.-я',
                      'Дата АПП', 'Дата ПАПП', 'Контрагент', 'Цена_дог,руб (дубли)',
                      'Площадь общая по договору (дубли)', 'Цена_дог,руб(без дублей)',
                      'Площадь общая по договору (без дублей)', 'Адрес дома', 'Застройка', 'Очередь', 'Помещение Тип',
                      'Помещение Под Тип',
                      'Помещение Студ', 'Корпус Номер', 'Этаж', '№кв.', 'Площадь общая по обмерам БТИ', '# пом.',
                      'Помещение', 'Правообладатель Название',
                      'Прав.Тип', '## плт. граф.', '### плт. факт', 'Дата_плт.', 'Сумма_плт., руб (без дублей)',
                      'График Платежей Задолженность (без дублей)',
                      'График Платежей Дней Просрочки', 'Дата платежа факт', 'Сумма, руб']

ROMANIAN_NUMBERS = pd.DataFrame({1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI',  7: 'VII',
                    8: 'VIII', 9: 'IX', 10: 'X', 11: 'XI', 12: 'XII' }.items(), columns = ['latin', 'romanian'])
d = {
    'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок': ((112, 48, 160), (255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Корректировка кв.м.': ((112, 48, 160), (255, 255, 255), 'Корректировка продаж по ДДУ'),
    'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок': ((112, 48, 160), (255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок': ((82, 170, 49),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Корректировка тыс.руб.': ((82, 170, 49),(255, 255, 255), 'Корректировка продаж по ДДУ'),
    'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок': ((82, 170, 49),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Партнерские продажи, кв. м.': ((4, 65, 149),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Партнерские продажи, тыс. руб.': ((4, 65, 149),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
}

period_dict = {'Месяц': [period for element in [[f'1_{year}', f'2_{year}', f'3_{year}', f'4_{year}', f'5_{year}', f'6_{year}', f'7_{year}', f'8_{year}', f'9_{year}', f'10_{year}', f'11_{year}', f'12_{year}']
                         for year in range(2014, 2020)] for period in element],
               'Квартал': [period for element in [[f'1_{year}', f'2_{year}', f'3_{year}', f'4_{year}']
                         for year in range(2014, 2020)] for period in element],
               'Полугодие': [period for element in [[f'1_{year}', f'2_{year}']
                         for year in range(2014, 2020)] for period in element],
               'Год': [str(year) for year in range(2014, 2020)]
               }

def create_bottom_dict(user_period):
    bottom_dict = {
        'По данным 1С': ((112, 48, 160), (255, 255, 255), period_dict[user_period], {
            'Реализация по ДДУ': ['Заключение', 'Расторжение', 'Поступление ДС', 'Возврат ДС', 'Прочее \nдвижение'],
            'Реализация по ДКП': ['Заключение', 'Расторжение']},
                         {
                             'Index 1': ['ДДУ', 'ДКП'],
                             'Index 2': ['Реализация_ДДУ', 'Реализация_ДКП'],
                             'Index 3': period_dict[user_period]
                         }),
        'По данным CRM': ((82, 170, 49), (255, 255, 255), period_dict[user_period], {
            'Реализация по ДДУ': ["Заключение \n(тыс.руб.)", "Расторжение \n(тыс.руб.)", "Заключение \n(кв.м.)",
                                  "Расторжение \n(кв.м.)"],
            'Реализация по ДКП': ["Заключение \n(тыс.руб.)", "Расторжение \n(тыс.руб.)", "Заключение \n(кв.м.)",
                                  "Расторжение \n(кв.м.)"]},
                          {
                              'Index 1': ['Реализация_ДДУ', 'Реализация_ДКП'],
                              'Index 2': period_dict[user_period],
                          }),
        'Разница': ((4, 65, 149), (255, 255, 255), period_dict[user_period], {'Разница': ['Заключение', 'Расторжение'],
                                                            'Разница по ДКП': ['Заключение', 'Расторжение']},
                    {
                        'Index 1': period_dict[user_period]
                    }),
    }
    return bottom_dict
def create_excel_file(path, name):
    Excel = win32com.client.Dispatch("Excel.Application")
    Excel.DisplayAlerts = False
    Excel.Visible = False
    try:
        wb = Excel.Workbooks.Add()
        path = os.path.abspath((path+name))
        wb.SaveAs(f'{path}', 50)
    except Exception as exp:
        Excel.Quit()
        sys.exit()
    else:
        wb.Save()
        wb.Close()
        Excel.Quit()
        return path

def open_and_fill_new_file(path, name, prj, df, period):
    filename = create_excel_file(path, name)
    Excel = win32com.client.Dispatch("Excel.Application")
    Excel.DisplayAlerts = False
    Excel.Visible = False
    wb = Excel.Workbooks.Open(filename)
    if Excel.ReferenceStyle != 1:
        Excel.ReferenceStyle = 1

    # sheet = wb.Worksheets('Лист1')
    # sheet.Cells(2.1).Value = "win32com"
    prj_for_formula = prj.replace('-', '_').replace(' ', '_')
    DDU_sheet, DKP_sheet, CRM_sheet, RES_sheet, DICT_sheet = create_sheet(wb, prj_for_formula)
    ws = wb.Worksheets(DICT_sheet)
    ws.Range(ws.Cells(1, 1),  # Cell to start the "paste"
             ws.Cells(1 + len(ROMANIAN_NUMBERS.index) - 1,
                      1 + len(ROMANIAN_NUMBERS.columns) - 1)  # No -1 for the index
             ).Value = ROMANIAN_NUMBERS.values
    sheet_data = namedtuple('SheetData', ['columns', 'add_columns', 'head', 'init_row',
                                          'init_col', 'interior_color', 'font_color', 'width', 'height'])
    wb.Worksheets(1).Delete()
    sheet_dict = {
        DDU_sheet: sheet_data(BIT_COLUMNS, ADD_COLUMNS, 'Карточка счета 76.33 (1С)', 5, 2, (112, 48, 160), (255, 255, 255), 10, 45),
        DKP_sheet: sheet_data(BIT_COLUMNS, ADD_COLUMNS, 'Карточка счета 90.01.1 (1С)', 5, 2, (112, 48, 160), (255, 255, 255), 10, 60),
        CRM_sheet: sheet_data(CRM_COLUMNS, ADD_CRM_COLUMNS, '', 4, 1, (255, 255, 255), (0, 0, 0), 13, 90)
    }
    create_columns(Excel, wb,  sheet_dict)
    ws = wb.Worksheets(RES_sheet)
    ws.Activate()
    Excel.ActiveWindow.DisplayGridlines = False
    ws.Cells.Font.Name = "Arial"
    ws.Cells.Font.Size = 10
    fill_data(ws, df, DDU_sheet, DKP_sheet, CRM_sheet,DICT_sheet, prj_for_formula, period, prj)
    ws = wb.Worksheets(RES_sheet)
    ws.Activate()
    wb.Worksheets(RES_sheet).Tab.Color = rgbToInt((128, 128, 128))
    wb.Worksheets(DDU_sheet).Tab.Color = rgbToInt((4, 65, 149))
    wb.Worksheets(DKP_sheet).Tab.Color = rgbToInt((112, 48, 160))
    wb.Worksheets(CRM_sheet).Tab.Color = rgbToInt((82, 170, 49))
    wb.Save()
    wb.Close()
    Excel.Quit()


def create_sheet(wb, prj):
    # удалить старый лист
    # передавать еще список имен или генерировать его
    sheetnames = ['Свод', 'ДДУ', "ДКП", "CRM", 'Словарь']
    for name in sheetnames:
        add = wb.Sheets.Add(Before=None, After=wb.Sheets(wb.Sheets.count))
        if name == 'Словарь':
            add.Name = name + '_' + prj.replace('-', '_').replace(" ", "_")
        else:
            add.Name = name + '_' + prj.replace(" ", "_")
    return (f'ДДУ_{prj.replace(" ", "_")}', f'ДКП_{prj.replace(" ", "_")}', f'CRM_{prj.replace(" ", "_")}', f'СВОД_{prj.replace(" ", "_")}', f'Словарь_{prj.replace("-", "_").replace(" ", "_")}')



def create_columns(Excel, wb, sheet_dict):

    for key, value in sheet_dict.items():

        ws = wb.Worksheets(key)
        ws.Activate()
        Excel.ActiveWindow.DisplayGridlines = False
        ws.Rows(value.init_row).RowHeight = value.height
        for i, element in enumerate(value.columns):
            ws.Columns(value.init_col+i).ColumnWidth = value.width
            ws.Cells(value.init_row, value.init_col+i).Value = element
            ws.Cells(value.init_row, value.init_col+i).WrapText = True
            align_cells(ws.Cells(value.init_row, value.init_col + i), value.interior_color, value.font_color, False)
        if 'CRM' not in key:
            rng = ws.Range(ws.Cells(value.init_row - 1, value.init_col), ws.Cells(value.init_row, value.init_col+i))
        else:
            rng = ws.Range(ws.Cells(value.init_row, value.init_col), ws.Cells(value.init_row, value.init_col + i))
        for border_id in range(7, 13):
            if border_id in (11, 12) and 'CRM' not in key:
                rng.Borders(border_id).LineStyle = -4119
                rng.Borders(border_id).Weight = 1
            else:
                rng.Borders(border_id).LineStyle = 1
                rng.Borders(border_id).Weight = 2
        ws.Cells(value.init_row - 1, value.init_col).Value = value.head
        align_cells(ws.Range(ws.Cells(value.init_row - 1, value.init_col), ws.Cells(value.init_row - 1, value.init_col+i)),
                    value.interior_color, value.font_color)
        for j, element in enumerate(value.add_columns, i + 2):
            ws.Columns(value.init_col + j).ColumnWidth = value.width
            ws.Cells(value.init_row, value.init_col + j).Value = element
            ws.Cells(value.init_row, value.init_col + j).WrapText = True
            align_cells(ws.Cells(value.init_row, value.init_col + j), rgbToInt((82, 170, 49)), rgbToInt((255, 255, 255)), False)

        rng = ws.Range(ws.Cells(value.init_row, value.init_col+i+2), ws.Cells(value.init_row, value.init_col+j))
        for border_id in range(7, 13):
            if border_id in (11, 12):
                rng.Borders(border_id).LineStyle = -4119
                rng.Borders(border_id).Weight = 1
            else:
                rng.Borders(border_id).LineStyle = 1
                rng.Borders(border_id).Weight = 2

def fill_data(ws,df, DDU_name, DKP_name, CRM_name, DICT_name,  prj, period, prj_for_data):
    bottom_dict = create_bottom_dict(period)
    periods = period_dict[period]
    ws.Cells(19,2).Value = "Данные по продажам"
    decorate_cells(ws.Range(ws.Cells(19,2), ws.Cells(19,4+len(periods))), rgbToInt((57, 115, 179)), rgbToInt((255, 255, 255)), 30, 'Arial', True)
    decorate_cells(ws.Range(ws.Cells(19, 4+len(periods)+ 2),ws.Cells(19, 4+len(periods)*2+ 1)), rgbToInt((57, 115, 179)), rgbToInt((255, 255, 255)), 30,
                   'Arial', True)

    StartRow = 47 + len(d) + 2*len(d) + 4*len(d) + len(d)*len(df)

    ws.Cells(StartRow, 2).Value = "СВЕРКА 1С и CRM"
    ws.Cells(StartRow, 2).Interior.Color = rgbToInt((255, 255, 255))
    ws.Cells(StartRow, 2).Font.Color = rgbToInt((0, 0, 0))
    ws.Cells(StartRow, 2).Font.Size = 30
    ws.Cells(StartRow, 2).Font.Name = 'Arial'
    ws.Cells(StartRow, 2).Font.Bold = True
    ws.Cells(StartRow, 2).Font.Italic = True
    StartRow+=2
    subhead_font_color = (0, 0, 0)
    subhead_interior_color = (255, 255, 255)
    BIT_row, CRM_row = 0, 0
    for k, v in bottom_dict.items():
        ws.Cells(StartRow, 2).Value = k
        decorate_cells(ws.Range(ws.Cells(StartRow, 2), ws.Cells(StartRow, 5)), rgbToInt(subhead_interior_color),
                       rgbToInt(subhead_font_color), 20, 'Arial', True)
        index_dict = create_index_dict(k, periods)
        try:
            index_frame = pd.DataFrame().from_dict(index_dict, orient='columns')
        except Exception as exp:
            index_frame = pd.DataFrame([index_dict])
        if k == 'По данным 1С':
            BIT_row = StartRow + len(index_frame) + 4
        elif k == 'По данным CRM':
            CRM_row = StartRow + len(index_frame) + 4
        column_dict = create_column_dict(k, periods, StartRow, DICT_name)
        #     pattern1 = ['ДДУ', 'ДДУ', 'ДДУ', 'ДДУ', 'ДДУ', '', 'ДКП', 'ДКП', '', '']
        #     pattern2 = ['Реализация_ДДУ', 'Реализация_ДДУ', 'ДДУ', 'ДДУ', '', '', 'Реализация_ДКП', 'Реализация_ДКП',
        #                 'ДКП', 'ДКП']
        #     count = 0
        #     dct = dict()
        #     for period in periods:
        #         pattern3 = [period] * 10
        #         for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
        #             if count == 0:
        #                 dct[str(-1)] = ['Index 1', 'Index 2', 'Index 3']
        #             dct[str(count)] = [ptrn1, ptrn2, ptrn3]
        #             count += 1
        #
        #     pattern1 = []
        #     column_dict = dict()
        #     count = 0
        #     pattern2 = ['Реализация по ДДУ', '', '', '', '', '', 'Реализация по ДКП', '', '', '']
        #     pattern3 = ['Заключение', 'Расторжение', 'Поступление ДС', 'Возврат ДС', 'Прочее движение', '',
        #                     'Заключение', 'Расторжение', '', '']
        #     for i in range(len(periods)):
        #         formula = f'=IF(ISERROR(SEARCH("_",{get_column_letter(5+10*i)}{StartRow + 3},1)),' \
        #                   f'CONCATENATE({get_column_letter(5+10*i)}{StartRow + 3}," год"),' \
        #                   f'CONCATENATE(VLOOKUP(MID({get_column_letter(5+10*i)}{StartRow + 3},' \
        #                   f'1,SEARCH("_",{get_column_letter(5+10*i)}{StartRow + 3})-1)+0,' \
        #                   f'{DICT_name}!$A:$B,2,0),"-е ",' \
        #                   f'RIGHT({get_column_letter(5+10*i)}{StartRow + 3},4)," года"))'
        #         pattern1.append(formula)
        #         pattern1.extend(['', '', '', '', '', '', '', '', ''])
        #
        #         for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
        #             column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
        #             count += 1
        #         pattern1 = []
        #
        #
        #     for _ in periods:
        #         for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
        #             column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
        #             count += 1
        #     DDU_FORMULA = f'=SUMIFS({DDU_name}!$AB:$AB,' \
        #                   f'{DDU_name}!$X:$X,${StartRow + 6}:${StartRow + 6},' \
        #                   f'{DDU_name}!$U:$U,${StartRow + 3}:${StartRow + 3},' \
        #                   f'{DDU_name}!$V:$V,$C:$C,{DDU_name}!$W:$W,$D:$D)'
        #     DKP_FORMULA = f'=SUMIFS({DKP_name}!$AB:$AB,' \
        #                   f'{DKP_name}!$X:$X,${StartRow + 6}:${StartRow + 6},' \
        #                   f'{DKP_name}!$U:$U,${StartRow + 3}:${StartRow + 3},' \
        #                   f'{DKP_name}!$V:$V,$C:$C,{DKP_name}!$W:$W,$D:$D)'
        #     value_dict = dict()
        #     count = 5
        #     for _ in periods:
        #         for i in range(10):
        #             if i <=4:
        #                 lst = [DDU_FORMULA]*len(df)
        #             elif i == 5 or i in (8, 9):
        #                 lst = [''] * len(df)
        #             else:
        #                 lst = [DKP_FORMULA] * len(df)
        #             lst.extend(['', f'=SUM({get_column_letter(count)}{StartRow + 5}:{get_column_letter(count)}{StartRow + 5 + len(df) - 1})'])
        #             value_dict[str(count)] = lst
        #             count+=1
        #
        # elif k == 'По данным CRM':
        #     CRM_row = StartRow + 6
        #     pattern1 = ['Реализация_ДДУ', 'Реализация_ДДУ', 'ДДУ', 'ДДУ', '', '', 'Реализация_ДКП', 'Реализация_ДКП',
        #                 'ДКП', 'ДКП']
        #     count = 0
        #     dct = dict()
        #     for period in periods:
        #         pattern2 = [period] * 10
        #         for ptrn1, ptrn2 in zip(pattern1, pattern2):
        #             if count == 0:
        #                 dct[str(-1)] = ['Index 1', 'Index 2']
        #             dct[str(count)] = [ptrn1, ptrn2]
        #             count += 1
        #     pattern2 = ['Реализация по ДДУ', '', '', '', '', '', 'Реализация по ДКП', '','', '']
        #     pattern3 = ['Заключение\n(тыс.руб.)', 'Расторжение\n(тыс.руб.)', 'Заключение\n(кв.м.)', 'Расторжение\n(кв.м.)', '',
        #                 '', 'Заключение\n(тыс.руб.)', 'Расторжение\n(тыс.руб.)', 'Заключение\n(кв.м.)', 'Расторжение\n(кв.м.)']
        #     column_dict = dict()
        #     for i in range(len(periods)):
        #         pattern1 = []
        #         formula = f'=IF(ISERROR(SEARCH("_",{get_column_letter(5+10*i)}{StartRow + 2},1)),' \
        #                   f'CONCATENATE({get_column_letter(5+10*i)}{StartRow + 2}," год"),' \
        #                   f'CONCATENATE(VLOOKUP(MID({get_column_letter(5+10*i)}{StartRow + 2},' \
        #                   f'1,SEARCH("_",{get_column_letter(5+10*i)}{StartRow + 2})-1)+0,' \
        #                   f'{DICT_name}!$A:$B,2,0),"-е ",' \
        #                   f'RIGHT({get_column_letter(5+10*i)}{StartRow + 2},4)," года"))'
        #         pattern1.append(formula)
        #         pattern1.extend(['', '', '', '', '', '', '', '', ''])
        #         for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
        #             column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
        #             count += 1
        #
        #     DDU_FORMULA_METRES_CLOSE = FORMULA_METRES_CLOSE.substitute(CRM_NAME = CRM_name, SALES_PERIOD = StartRow + 2, DOC_TYPE = 'ДДУ')
        #     DKP_FORMULA_METRES_CLOSE = FORMULA_METRES_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=StartRow + 2,DOC_TYPE='ДКП')
        #     DDU_FORMULA_METRES_DENIAL = FORMULA_METRES_DENIAL.substitute(CRM_NAME = CRM_name, SALES_PERIOD = StartRow + 2, DOC_TYPE = 'ДДУ')
        #     DKP_FORMULA_METRES_DENIAL = FORMULA_METRES_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=StartRow + 2,DOC_TYPE='ДКП')
        #     DDU_FORMULA_MONEY_CLOSE = FORMULA_MONEY_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=StartRow + 2,DOC_TYPE='ДДУ')
        #     DKP_FORMULA_MONEY_CLOSE = FORMULA_MONEY_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=StartRow + 2, DOC_TYPE='ДКП')
        #     DDU_FORMULA_MONEY_DENIAL = FORMULA_MONEY_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=StartRow + 2,DOC_TYPE='ДДУ')
        #     DKP_FORMULA_MONEY_DENIAL = FORMULA_MONEY_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=StartRow + 2,
        #                                                                DOC_TYPE='ДКП')
        #
        #     FORMULS = [DDU_FORMULA_METRES_CLOSE, DDU_FORMULA_METRES_DENIAL,DDU_FORMULA_MONEY_CLOSE, DDU_FORMULA_MONEY_DENIAL, '', '',
        #                DKP_FORMULA_METRES_CLOSE, DKP_FORMULA_METRES_DENIAL,DKP_FORMULA_MONEY_CLOSE, DKP_FORMULA_MONEY_DENIAL ]
        #     value_dict = dict()
        #     count = 5
        #     for _ in periods:
        #         for i, formula in enumerate(FORMULS):
        #             lst = [formula] * len(df)
        #             lst.extend(['', f'=SUM({get_column_letter(count)}{StartRow + 5}:{get_column_letter(count)}{StartRow + 5 + len(df) - 1})'])
        #             value_dict[str(count)] = lst
        #             count+=1
        #
        #
        # else:
        #     count = 0
        #     dct = dict()
        #     for period in periods:
        #         pattern1 = [period] * 10
        #         for ptrn1 in pattern1:
        #             if count == 0:
        #                 dct[str(-1)] = 'Index 1'
        #             dct[str(count)] = ptrn1
        #             count += 1
        #
        #     pattern2 = ['Разница', '', '', '', '', '', '', 'Разница по ДКП', '','']
        #     pattern3 = ['Заключение', 'Расторжение', '', '', '', '',
        #                 'Заключение', 'Расторжение','', '',]
        #     column_dict = dict()
        #     for i in range(len(periods)):
        #         pattern1 = []
        #         formula = f'=IF(ISERROR(SEARCH("_",{get_column_letter(5+10*i)}{StartRow + 1},1)),' \
        #                   f'CONCATENATE({get_column_letter(5+10*i)}{StartRow + 1}," год"),' \
        #                   f'CONCATENATE(VLOOKUP(MID({get_column_letter(5+10*i)}{StartRow + 1},' \
        #                   f'1,SEARCH("_",{get_column_letter(5+10*i)}{StartRow + 1})-1)+0,' \
        #                   f'{DICT_name}!$A:$B,2,0),"-е ",' \
        #                   f'RIGHT({get_column_letter(5+10*i)}{StartRow + 1},4)," года"))'
        #         pattern1.append(formula)
        #         pattern1.extend(['', '', '', '', '', '', '', '', ''])
        #         for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
        #             column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
        #             count += 1
        #     # for _ in periods:
        #     #     for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
        #     #         column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
        #     #         count += 1
        #     value_dict = dict()
        #     count = 5
        #     for _ in periods:
        #         for i in range(10):
        #             if i in(0, 1, 6, 7):
        #                 lst =  [f'={BIT_row + k}:{BIT_row + k}-{CRM_row + k}:{CRM_row + k}' for k in range(len(df))]
        #                 # value_dict[str(count)] = [f'{BIT_row}:{BIT_row}-{CRM_row}:{CRM_row}'] * 10
        #             else:
        #                 lst = [''] *  len(df)
        #                 # value_dict[str(count)] = [''] * 10
        #             lst.extend(['',
        #                         f'=SUM({get_column_letter(count)}{StartRow + 5}:{get_column_letter(count)}{StartRow + 5 + len(df) - 1})'])
        #             value_dict[str(count)] = lst
        #             count+=1

        StartRow += 1
        ws.Range(ws.Cells(StartRow, 4),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(index_frame.index) - 1,
                          4 + len(index_frame.columns) - 1)  # No -1 for the index
                 ).Value = index_frame.values
        align_cells(ws.Range(ws.Cells(StartRow, 4), ws.Cells(StartRow + len(index_frame.index) - 1, 4 + len(index_frame.columns) - 1)),
                    rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                    False)
        # добавить кол-во отступов
        StartRow += len(index_frame)
        for i, subhead in enumerate(['Компания', 'Очередь', 'Дом'], 2):
            ws.Cells(StartRow, i).Value = subhead
            MergeRange = ws.Range(ws.Cells(StartRow, i), ws.Cells(StartRow + 2, i))
            align_cells(MergeRange, v[0], v[1])

        column_frame = pd.DataFrame().from_dict(column_dict, orient='columns')
        # StartRow += 1
        column_rng = ws.Range(ws.Cells(StartRow, 5),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(column_frame.index) - 1,
                          5 + len(column_frame.columns) - 1))
        column_rng.Value = column_frame.values
        decorate_cells(column_rng, v[0], v[1], 10, 'Arial', True)
        StartRow+=3
        first_row  = {'Проект': 'Если необходимо вставить дополнительные очереди/дома, то добавьте строку перед текущей', 'Очередь': '', 'Дом': ''}
        second_row = {'Проект': 'ИТОГО', 'Очередь': '', 'Дом': ''}
        new_df = pd.concat([df, pd.DataFrame(first_row, index=[0]), pd.DataFrame(second_row, index=[0])])
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(new_df.index) - 1,
                          2 + len(new_df.columns) - 1)  # No -1 for the index
                 ).NumberFormat = '@'
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                    ws.Cells(StartRow + len(new_df.index) - 1,
                                2 + len(new_df.columns) - 1)  # No -1 for the index
                    ).Value = new_df.values


        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(new_df.index) - 1,
                          2 + len(new_df.columns) - 1)  # No -1 for the index
                 ).Replace(',', '.')

        value_dict = create_value_dict(ws, df, k, prj_for_data, periods, StartRow, BIT_row, CRM_row, DDU_name, DKP_name, CRM_name)
        values_frame = pd.DataFrame().from_dict(value_dict, orient='columns')
        # StartRow += 1
        ws.Range(ws.Cells(StartRow, 5),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(values_frame.index) - 1,
                          5 + len(values_frame.columns) - 1)  # No -1 for the index
                 ).Value = values_frame.values
        # set_border(ws, rng, StartRow - 3, 2, StartRow + len(df) + 1, 10 * length + 4, rgbToInt(v[0]), True)
        rng = ws.Range(ws.Cells(StartRow, 5), ws.Cells(StartRow + len(values_frame.index) - 1,
                          5 + len(values_frame.columns) - 1))
        rng.NumberFormat = '# ###;(# ###);"-"'
        rng = ws.Range(ws.Cells(StartRow-3, 2), ws.Cells(StartRow + len(values_frame.index) - 1,
                                                       5 + len(values_frame.columns) - 1))
        set_border(ws, rng, StartRow - 3, 2, StartRow + len(new_df)-1, 4+10*len(periods), rgbToInt(v[0]), True)

        if k == 'Разница':
            pass
        for i in range(len(periods)):
            new_align_cells(ws.Range(ws.Cells(StartRow - 3, 5 + 10 * i),
                     ws.Cells(StartRow - 3, 14 + 10 * i)))

            new_align_cells(ws.Range(ws.Cells(StartRow - 2, 5 + 10 * i),
             ws.Cells(StartRow - 2, 10 + 10 * i)))

            new_align_cells(ws.Range(ws.Cells(StartRow - 2, 11 + 10 * i),
             ws.Cells(StartRow - 2, 14 + 10 * i)))

            if k == 'По данным CRM':
                rng = ws.Range(ws.Cells(StartRow-1, 9 + 10*i),  # Cell to start the "paste"
                     ws.Cells(StartRow + len(values_frame.index) - 3,10 + 10*i))
                fill_pattern(rng)
            elif k == 'По данным 1С':
                rng = ws.Range(ws.Cells(StartRow-1, 10 + 10*i),  # Cell to start the "paste"
                     ws.Cells(StartRow + len(values_frame.index) - 3,10 + 10*i))
                fill_pattern(rng)
                rng = ws.Range(ws.Cells(StartRow-1, 13 + 10*i),  # Cell to start the "paste"
                     ws.Cells(StartRow + len(values_frame.index) - 3,14 + 10*i))
                fill_pattern(rng)
            else:
                rng = ws.Range(ws.Cells(StartRow - 1, 7 + 10 * i),  # Cell to start the "paste"
                               ws.Cells(StartRow + len(values_frame.index) - 3, 10 + 10 * i))
                fill_pattern(rng)
                rng = ws.Range(ws.Cells(StartRow - 1, 13 + 10 * i),  # Cell to start the "paste"
                               ws.Cells(StartRow + len(values_frame.index) - 3, 14 + 10 * i))
                fill_pattern(rng)

        StartRow+=len(df) + 4

    StartRow = 21
    for k, v in d.items():
        ws.Cells(StartRow, 2).Value = k
        decorate_cells(ws.Range(ws.Cells(StartRow, 2), ws.Cells(StartRow, 4 + len(periods))), rgbToInt(v[0]), rgbToInt(v[1]), 20,
                           'Arial', True)
        decorate_cells(ws.Range(ws.Cells(StartRow, 4 + len(periods) + 2), ws.Cells(StartRow, 4 + len(periods)*2 + 1)), rgbToInt(v[0]), rgbToInt(v[1]), 20,
                       'Arial', True)
        StartRow += 1
        period_df = pd.DataFrame([{str(i): period for i, period in enumerate(periods)}])
        if k  == 'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок':
            CORRECTED_METRES_ROW, CORRECTED_MONEY_ROW = 1, 1
            Sales_metres_row = StartRow + 1
            ws.Range(ws.Cells(StartRow + 1, 5), ws.Cells(StartRow + 1, 5 + len(period_df.columns) - 1)).Value = period_df.values
            ws.Range(ws.Cells(StartRow + 1, 6 + len(periods)),ws.Cells(StartRow + 1, 6 + len(periods) + len(period_df.columns) - 1)).Value = period_df.values
            align_cells(ws.Range(ws.Cells(StartRow + 1, 5), ws.Cells(StartRow + 1, 5 + len(period_df.columns) - 1)),
                        rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
            align_cells(ws.Range(ws.Cells(StartRow + 1, 6 + len(periods)),ws.Cells(StartRow + 1, 6 + len(periods) + len(period_df.columns) - 1)),
                        rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
        elif k == 'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок':
            ws.Range(ws.Cells(StartRow + 1, 5), ws.Cells(StartRow + 1, 5 + len(period_df.columns) - 1)).Value = period_df.values
            ws.Range(ws.Cells(StartRow + 1, 6 + len(periods)), ws.Cells(StartRow + 1, 6 + len(periods) + len(period_df.columns) - 1)).Value = period_df.values
            align_cells(ws.Range(ws.Cells(StartRow + 1, 5), ws.Cells(StartRow + 1, 5 + len(period_df.columns) - 1)),
                        rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
            align_cells(ws.Range(ws.Cells(StartRow + 1, 6 + len(periods)),
                                 ws.Cells(StartRow + 1, 6 + len(periods) + len(period_df.columns) - 1)),
                        rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
        StartRow += 2
        for i, subhead in enumerate(['Компания', 'Очередь', 'Дом'], 2):
            ws.Cells(StartRow, i).Value = subhead
            MergeRange = ws.Range(ws.Cells(StartRow, i), ws.Cells(StartRow + 3, i))
            align_cells(MergeRange, v[0], v[1])

        ws.Cells(StartRow, 5).Value = v[2]
        ws.Range(ws.Cells(StartRow, 5), ws.Cells(StartRow, 5+len(periods))).HorizontalAlignment = 7
        decorate_cells(ws.Range(ws.Cells(StartRow, 5), ws.Cells(StartRow, 5+len(periods) - 1)), v[0], v[1], 10, 'Arial', True)
        ws.Cells(StartRow, 6+len(periods)).Value = v[2]
        ws.Range(ws.Cells(StartRow,  6+len(periods)), ws.Cells(StartRow,  6+2*len(periods))).HorizontalAlignment = 7
        decorate_cells(ws.Range(ws.Cells(StartRow,  6+len(periods)), ws.Cells(StartRow,  5+2*len(periods))), v[0], v[1],10, 'Arial', True)

        column_df = pd.DataFrame([{str(i):  f'=IF(ISERROR(SEARCH("_",{get_column_letter(i)}{Sales_metres_row},1)),' \
        f'CONCATENATE({get_column_letter(i)}{Sales_metres_row}," год"),CONCATENATE(VLOOKUP(MID({get_column_letter(i)}{Sales_metres_row},' \
        f'1,SEARCH("_",{get_column_letter(i)}{Sales_metres_row})-1)+0,{DICT_name}!$A:$B,2,0),"-е ",' \
        f'RIGHT({get_column_letter(i)}{Sales_metres_row},4)," года"))' for i, period in enumerate(periods, 5) }])
        ws.Range(ws.Cells(StartRow+1, 5), ws.Cells(StartRow+1, 5 + len(column_df.columns) - 1)).Value = column_df.values

        column_df = pd.DataFrame([{str(i):  f'=IF(ISERROR(SEARCH("_",{get_column_letter(i)}{Sales_metres_row},1)),' \
        f'CONCATENATE({get_column_letter(i)}{Sales_metres_row}," год"),CONCATENATE(VLOOKUP(MID({get_column_letter(i)}{Sales_metres_row},' \
        f'1,SEARCH("_",{get_column_letter(i)}{Sales_metres_row})-1)+0,{DICT_name}!$A:$B,2,0),"-е ",' \
        f'RIGHT({get_column_letter(i)}{Sales_metres_row},4)," года"))' for i, period in enumerate(periods, 6+len(periods)) }])
        ws.Range(ws.Cells(StartRow+1, 6+len(periods)), ws.Cells(StartRow+1, 6+len(periods) + len(column_df.columns) - 1)).Value = column_df.values
        StartRow += 4
        first_row = {'Проект': '',
                     'Очередь': '', 'Дом': ''}
        second_row = {'Проект': 'ИТОГО', 'Очередь': '', 'Дом': ''}
        new_df = pd.concat([df, pd.DataFrame(first_row, index=[0]), pd.DataFrame(second_row, index=[0])])
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(new_df.index) - 1,
                          2 + len(new_df.columns) - 1)  # No -1 for the index
                 ).NumberFormat = '@'
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(new_df.index) - 1,
                          2 + len(new_df.columns) - 1)  # No -1 for the index
                 ).Value = new_df.values
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(new_df.index) - 1,
                          2 + len(new_df.columns) - 1)  # No -1 for the index
                 ).Replace(',', '.')
        left_value_dict = dict()
        right_value_dict = dict()
        for j, period in enumerate(periods, 5):
            align_cells(ws.Range(ws.Cells(StartRow-3, j), ws.Cells(StartRow-1, j)), v[0], v[1])
            align_cells(ws.Range(ws.Cells(StartRow - 3, len(periods) + j + 1), ws.Cells(StartRow - 1, len(periods) + j + 1)), v[0], v[1])
            if k == 'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок':
                if j == 5:
                    left_lst = [SALES_METRES_WITH_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДДУ')
                            for i in range(len(df))]
                    right_lst = [SALES_METRES_WITH_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДКП')
                            for i in range(len(df))]
                else:
                    left_lst = [SALES_METRES_WITH_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДДУ', PREV_COL=get_column_letter(j-1))
                            for i in range(len(df))]
                    right_lst = [SALES_METRES_WITH_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДКП', PREV_COL=get_column_letter(j+len(periods)))
                            for i in range(len(df))]
            elif k == 'Корректировка кв.м.':
                CORRECTED_METRES_ROW = StartRow + 1
                left_lst = [CORRECTED_METRES.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row, DOC_TYPE='ДДУ')] * len(df)
                right_lst = [CORRECTED_METRES.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,
                                                        DOC_TYPE='ДКП')] * len(df)
            elif k == 'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок':
                if j == 5:
                    left_lst = [SALES_METRES_WITHOUT_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДДУ',SALES_CORRECTED=CORRECTED_METRES_ROW+i)
                                for i in range(len(df))]
                    right_lst = [SALES_METRES_WITHOUT_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДКП',SALES_CORRECTED=CORRECTED_METRES_ROW+i)
                                for i in range(len(df))]
                else:
                    left_lst = [SALES_METRES_WITHOUT_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДДУ',SALES_CORRECTED=CORRECTED_METRES_ROW+i, PREV_COL=get_column_letter(j - 1))
                                for i in range(len(df))]
                    right_lst = [SALES_METRES_WITHOUT_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДКП',SALES_CORRECTED=CORRECTED_METRES_ROW+i,PREV_COL=get_column_letter(j+len(periods)))
                                for i in range(len(df))]

            elif k == 'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок':
                if j == 5:
                    left_lst = [SALES_MONEY_WITH_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДДУ')
                            for i in range(len(df))]
                    right_lst = [SALES_MONEY_WITH_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДКП')
                            for i in range(len(df))]
                else:
                    left_lst = [SALES_MONEY_WITH_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДДУ', PREV_COL=get_column_letter(j-1))
                            for i in range(len(df))]
                    right_lst = [SALES_MONEY_WITH_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5, DOC_TYPE='ДКП', PREV_COL=get_column_letter(j+len(periods)))
                            for i in range(len(df))]
            elif k == 'Корректировка тыс.руб.':
                CORRECTED_MONEY_ROW = StartRow + 1
                left_lst = [CORRECTED_SALES.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,
                                                        DOC_TYPE='ДДУ')] * len(df)
                right_lst = [CORRECTED_SALES.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,
                                                         DOC_TYPE='ДКП')] * len(df)
            elif k == 'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок':
                if j == 5:
                    left_lst = [SALES_MONEY_WITHOUT_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДДУ',SALES_CORRECTED=CORRECTED_MONEY_ROW+i-1)
                                for i in range(len(df))]
                    right_lst = [SALES_MONEY_WITHOUT_FIRST_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДКП',SALES_CORRECTED=CORRECTED_MONEY_ROW+i-1)
                                for i in range(len(df))]
                else:
                    left_lst = [SALES_MONEY_WITHOUT_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДДУ',SALES_CORRECTED=CORRECTED_MONEY_ROW+i-1, PREV_COL=get_column_letter(j - 1))
                                for i in range(len(df))]
                    right_lst = [SALES_MONEY_WITHOUT_OTHER_COLUMN.substitute(CRM_ROW=CRM_row + i, CRM_PERIOD=CRM_row-4,SALES_PERIOD=Sales_metres_row,CRM_TYPE=CRM_row-5,DOC_TYPE='ДКП',SALES_CORRECTED=CORRECTED_MONEY_ROW+i-1,PREV_COL=get_column_letter(j+len(periods)))
                                for i in range(len(df))]
            elif k == 'Партнерские продажи, кв. м.':
                if j == 5:
                    left_lst = [PARTNER_SALES_METRES_FIRST_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДДУ')
                                for i in range(len(df))]
                    right_lst = [PARTNER_SALES_METRES_FIRST_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДКП')
                                for i in range(len(df))]
                else:
                    left_lst = [PARTNER_SALES_METRES_OTHER_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДДУ', PREV_COL=get_column_letter(j - 1))
                                for i in range(len(df))]
                    right_lst = [PARTNER_SALES_METRES_OTHER_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДКП',PREV_COL=get_column_letter(j+len(periods)))
                                for i in range(len(df))]
            else:
                if j == 5:
                    left_lst = [PARTNER_SALES_MONEY_FIRST_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДДУ')
                                for i in range(len(df))]
                    right_lst = [PARTNER_SALES_MONEY_FIRST_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДКП')
                                for i in range(len(df))]
                else:
                    left_lst = [PARTNER_SALES_MONEY_OTHER_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДДУ', PREV_COL=get_column_letter(j - 1))
                                for i in range(len(df))]
                    right_lst = [PARTNER_SALES_MONEY_OTHER_COLUMN.substitute(CRM_NAME=CRM_name, SALES_PERIOD=Sales_metres_row,DOC_TYPE='ДКП',PREV_COL=get_column_letter(j+len(periods)))
                                for i in range(len(df))]


            left_lst.extend(['', f'=SUM({get_column_letter(j)}{StartRow}:{get_column_letter(j)}{StartRow+len(df)-1})'])
            left_value_dict[str(j)] = left_lst

            right_lst.extend(['', f'=SUM({get_column_letter(6 + j-5 + len(periods))}{StartRow}:{get_column_letter(6 + j-5 + len(periods))}{StartRow + len(df) - 1})' ])
            right_value_dict[str(j)] = right_lst

        left_value_frame = pd.DataFrame().from_dict(left_value_dict, orient='columns')
        right_value_frame = pd.DataFrame().from_dict(right_value_dict, orient='columns')

        rng = ws.Range(ws.Cells(StartRow, 5), ws.Cells(StartRow + len(left_value_frame.index) - 1, 5 + len(left_value_frame.columns) - 1))
        rng.Value = left_value_frame.values
        rng = ws.Range(ws.Cells(StartRow, 2),ws.Cells(StartRow + len(left_value_frame.index) - 1, 5 + len(left_value_frame.columns) - 1))
        set_border(ws, rng, StartRow, 2, StartRow + len(left_value_frame.index) - 1, 5 + len(left_value_frame.columns) - 1)

        rng.NumberFormat = '# ###;(# ###);"-"'
        rng = ws.Range(ws.Cells(StartRow, 5), ws.Cells(StartRow + len(left_value_frame.index) - 3, 5 + len(left_value_frame.columns) - 1))
        rng.Interior.Color = rgbToInt((221, 235, 247))

        rng = ws.Range(ws.Cells(StartRow, 6+len(periods)), ws.Cells(StartRow + len(right_value_frame.index) - 1,  6+len(periods) + len(right_value_frame.columns) - 1))
        rng.Value = right_value_frame.values
        set_border(ws, rng, StartRow, 6+len(periods), StartRow + len(right_value_frame.index) - 1, 6+len(periods) + len(right_value_frame.columns) - 1, option = True)
        rng.NumberFormat = '# ###;(# ###);"-"'

        rng = ws.Range(ws.Cells(StartRow, 6 + len(periods)), ws.Cells(StartRow + len(right_value_frame.index) - 3, 6 + len(periods) + len(right_value_frame.columns) - 1))
        rng.Interior.Color = rgbToInt((221, 235, 247))
        StartRow += len(df) + 3

    ws.Cells(2, 2).Value = f'СЗ САМОЛЕТ - {prj}'
    decorate_cells(ws.Cells(2, 2), (255, 255, 255), (0, 0, 0), 20, 'Arial', True, True)
    ws.Cells(4, 2).Value = 'Проверка полноты отражения данных'
    decorate_cells(ws.Cells(4, 2), (255, 255, 255), (0, 0, 0), 15, 'Arial', True, False, True)
    ws.Cells(6, 2).Value = 'Сумма продаж, тыс.руб.'
    decorate_cells(ws.Cells(6, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', True)
    ws.Cells(8, 2).Value = 'ПРОДАЖИ'
    rng = ws.Range(ws.Cells(8, 2), ws.Cells(9, 2))
    align_cells(rng, (112, 48, 160), (255, 255, 255))
    ws.Cells(10, 2).Value = 'Сумма продаж по ДДУ'
    decorate_cells(ws.Cells(10, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', False)
    ws.Cells(11, 2).Value = 'Сумма продаж по ДКП'
    decorate_cells(ws.Cells(11, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', False)
    ws.Cells(12, 2).Value = 'ИТОГО'
    decorate_cells(ws.Cells(12, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', False)

    ws.Cells(8, 3).Value = 'СУММА ПО'
    rng = ws.Range(ws.Cells(8, 3),  ws.Cells(8, 5))
    align_cells(rng, (82, 170, 49),(255, 255, 255))
    ws.Cells(9, 3).Value = '1С'
    rng = ws.Range(ws.Cells(9, 3), ws.Cells(9, 5))
    align_cells(rng, (82, 170, 49),(255, 255, 255),  False)
    # decorate_cells(ws.Cells(9, 3), (82, 170, 49),(255, 255, 255), 10, 'Arial', True)
    ws.Cells(10,
             3).Value = f'=SUMIFS({DDU_name}!AB:AB,{DDU_name}!X:X,"Заключение")+SUMIFS({DDU_name}!AB:AB,{DDU_name}!X:X,"Расторжение")'
    ws.Cells(10, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11, 3).Value = f'=SUMIFS({DKP_name}!AB:AB,{DKP_name}!X:X,"Заключение")'
    ws.Cells(11, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(12, 3).Value = f'=SUM(C10:C11)'
    ws.Cells(12, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(9, 4).Value = 'CRM'
    # decorate_cells(ws.Cells(9, 4), (82, 170, 49), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10,
             4).Value = f'=SUMIFS({CRM_name}!AY:AY,{CRM_name}!AQ:AQ,"ДДУ")-SUMIFS({CRM_name}!AY:AY,{CRM_name}!BA:BA,"Да")'
    ws.Cells(10, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11, 4).Value = f'=SUMIFS({CRM_name}!AY:AY,{CRM_name}!AQ:AQ,"ДКП")'
    ws.Cells(11, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(12, 4).Value = f'=SUM(D10:D11)'
    ws.Cells(12, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(9, 5).Value = 'СВОД'
    # decorate_cells(ws.Cells(9, 5), (82, 170, 49), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 5).Value = f'=SUMIFS({BIT_row+len(df)+1}:{BIT_row+len(df)+1},{BIT_row-5}:{BIT_row-5},"Реализация_ДДУ")'
    ws.Cells(10, 5).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11,
             5).Value = f'=SUMIFS({BIT_row + len(df) + 1}:{BIT_row + len(df) + 1},{BIT_row - 5}:{BIT_row - 5},"Реализация_ДКП")'
    ws.Cells(11, 5).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(12, 5).Value = f'=SUM(E10:E11)'
    ws.Cells(12, 5).NumberFormat = '# ###;(# ###);"-"'

    ws.Cells(8, 6).Value = 'ПРОВЕРКА'
    rng = ws.Range(ws.Cells(8, 6), ws.Cells(8, 8))
    align_cells(rng, (4, 65, 149),(255, 255, 255))
    ws.Cells(9, 6).Value = '1С VS CRM'
    rng = ws.Range(ws.Cells(9, 6), ws.Cells(9, 8))
    align_cells(rng, (4, 65, 149),(255, 255, 255),  False)
    # decorate_cells(ws.Cells(9, 6), (4, 65, 149),(255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 6).Value = '=IF(ROUNDDOWN(C10,-3)=ROUNDDOWN(D10,-3),"КОРРЕКТНО","ОШИБКА")'

    ws.Cells(11, 6).Value = '=IF(ROUND(C11,1)=ROUND(D11,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(9, 7).Value = '1С VS СВОД'
    # decorate_cells(ws.Cells(9, 7), (4, 65, 149), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 7).Value = '=IF(ROUND(C10,1)=ROUND(E10,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(11, 7).Value = '=IF(ROUND(C11,1)=ROUND(E11,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(9, 8).Value = 'CRM VS СВОД'
    # decorate_cells(ws.Cells(9, 8), (4, 65, 149), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 8).Value = '=IF(ROUNDDOWN(D10,-3)=ROUNDDOWN(E10,-3),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(11, 8).Value = '=IF(ROUND(D11,1)=ROUND(E11,1),"КОРРЕКТНО","ОШИБКА")'
    rng = ws.Range(ws.Cells(10, 6), ws.Cells(11, 8))
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108

    ws.Cells(11, 9).Value = '=C11-D11'
    ws.Cells(11, 9).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11, 10).Value = 'n/s'
    ws.Cells(10, 9).Value = '=C10-D10'
    ws.Cells(10, 9).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(10, 10).Value = 'n/s'
    rng = ws.Range(ws.Cells(10, 9), ws.Cells(11, 10))
    decorate_cells(rng, (255, 255, 255), (255, 0, 0), 10, 'Arial', True, False, True)

    ws.Cells(14, 3).Value = '1С'
    ws.Cells(14, 4).Value = 'СВОД'
    ws.Cells(14, 5).Value = 'ПРОВЕРКА'
    rng = ws.Range(ws.Cells(14, 2), ws.Cells(14, 5))
    align_cells(rng, (82, 170, 49),(255, 255, 255),  False)
    ws.Cells(15, 2).Value = 'Конечное сальдо 76.33 (ДДУ)'
    ws.Cells(15, 3).Value = f'=SUM({DDU_name}!AB:AB)'
    ws.Cells(15, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(15, 4).Value = f'=SUMIFS({BIT_row+len(df)+1}:{BIT_row+len(df)+1},{BIT_row-6}:{BIT_row-6},"ДДУ")'
    ws.Cells(15, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(15, 5).Value = f'=IF(ROUND(C15,1)=ROUND(D15,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(16, 2).Value = 'Оборот 90.01.1 (ДКП)'
    ws.Cells(16, 3).Value = f'=SUMIFS({DKP_name}!AB:AB,{DKP_name}!X:X,"Заключение")'
    ws.Cells(16, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(16, 4).Value = f'=SUMIFS({BIT_row+len(df)+1}:{BIT_row+len(df)+1},{BIT_row-6}:{BIT_row-6},"ДКП")'
    ws.Cells(16, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(16, 5).Value = f'=IF(ROUND(C16,1)=ROUND(D16,1),"КОРРЕКТНО","ОШИБКА")'
    rng = ws.Range(ws.Cells(15, 5), ws.Cells(16, 5))
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108

    set_column_width(ws, periods)
    rng = ws.Range(ws.Cells(8, 2), ws.Cells(12, 8))
    set_border(ws, rng, 8, 2, 12, 8, option=False)

    ws.Range(ws.Cells(12, 2), ws.Cells(12, 8)).Borders(8).Weight = 3
    ws.Range(ws.Cells(8, 2), ws.Cells(12, 2)).Borders(10).Weight = 2
    ws.Range(ws.Cells(8, 3), ws.Cells(12, 5)).Borders(10).Weight = 2
    set_conditional_formatting(ws.Range(ws.Cells(10, 6), ws.Cells(11, 8)))
    set_conditional_formatting(ws.Range(ws.Cells(15, 5), ws.Cells(16, 5)))
    rng = ws.Range(ws.Cells(14, 2), ws.Cells(16, 5))
    set_column_width(ws, periods)
    for border_id in range(7, 13):
        if border_id in (11, 12):
            rng.Borders(border_id).LineStyle = -4119
            rng.Borders(border_id).Weight = 1
        else:
            rng.Borders(border_id).LineStyle = 1
            rng.Borders(border_id).Weight = 2
    EndRow = ws.Range("{0}{1}".format('B', ws.Rows.Count)).End(-4162).Row
    row = 1
    while EndRow >=row:
        ADDRESS =  ws.Range(f'B{row}:B{EndRow}').Find('ИТОГО')
        ADDRESS.Font.Bold = True
        if row> 1 and ws.Range(f'B{row - 1}:B{EndRow}').Find('Если необходимо вставить дополнительные очереди/дома. то добавьте строку перед текущей') != None:
            ws.Range(f'B{row - 1}:B{EndRow}').Find('Если необходимо вставить дополнительные очереди/дома. то добавьте строку перед текущей').Font.Color = rgbToInt((255, 0, 0))
            ws.Range(f'B{row - 1}:B{EndRow}').Find('Если необходимо вставить дополнительные очереди/дома. то добавьте строку перед текущей').Font.Bold = True
        row = ADDRESS.Row + 2


def set_conditional_formatting(rng):
    positive_cond = rng.FormatConditions.Add(1, 3, "КОРРЕКТНО")
    positive_cond.Interior.Color = rgbToInt((198, 239, 206))
    positive_cond.Font.Color = rgbToInt((0, 97, 0))
    negative_cond = rng.FormatConditions.Add(1, 3, "ОШИБКА")
    negative_cond.Interior.Color = rgbToInt((255, 199, 206))
    negative_cond.Font.Color = rgbToInt((156, 0, 6))
def select_count_col(headname):
    if headname in ('Реализация по ДДУ', 'Разница'):
        return 5
    else:
        return 3
def rgbToInt(rgb):
    if isinstance(rgb, tuple):
        colorInt = rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)
        return colorInt
    else:
        return rgb


def decorate_cells(rng, interor_color, font_color, size, font_name, is_bold, is_underline =False, is_italic = False):
    rng.Interior.Color = rgbToInt(interor_color)
    rng.Font.Color = rgbToInt(font_color)
    rng.Font.Size = size
    rng.Font.Name = font_name
    rng.Font.Bold = is_bold
    rng.Font.Underline = is_underline
    rng.Font.Italic = is_italic
def align_cells(rng, interior_color, font_color, is_merge = True):
    rng.MergeCells = is_merge
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108
    decorate_cells(rng, rgbToInt(interior_color), rgbToInt(font_color), 10, 'Arial', True)

def set_column_width(ws, periods):
    for i in range(2, 10*len(periods)):
        if i == 2:
            ws.Columns(i).ColumnWidth = 29
        else:
            ws.Columns(i).ColumnWidth = 15

def set_border(ws, rng, init_row, init_col, last_row, last_col, color = '', option = True):
    for border_id in range(7, 13):
        if border_id in (11, 12):
            rng.Borders(border_id).LineStyle = -4119
            rng.Borders(border_id).Weight = 1
        else:
            rng.Borders(border_id).LineStyle = 1
            if border_id == 9:
                rng.Borders(border_id).Weight = 3
            else:
                rng.Borders(border_id).Weight = 2

    ws.Range(ws.Cells(init_row, init_col), ws.Cells(last_row, init_col)).Borders(8).Weight = 2
    if option:
        ws.Range((ws.Cells(last_row, init_col)), (ws.Cells(last_row, last_col))).Borders(8).Weight = 3
        ws.Range((ws.Cells(last_row-1, init_col)), (ws.Cells(last_row-1, last_col))).Borders(8).Weight = 2
        ws.Range((ws.Cells(last_row-1, init_col)), (ws.Cells(last_row-1, last_col))).Interior.Color = rgbToInt((112, 48, 160))
    else:
        ws.Range((ws.Cells(last_row, init_col)), (ws.Cells(last_row, last_col))).Borders(8).Weight = 3
        ws.Range(ws.Cells(init_row, init_col+1), ws.Cells(last_row, init_col+3)).Borders(10).Weight = 2

def new_align_cells(rng, is_merge = True):
    rng.MergeCells = is_merge
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108
    # decorate_cells(rng, rgbToInt(interior_color), rgbToInt(font_color), 10, 'Arial', True)


def fill_pattern(rng):
    # rng.Interior.Pattern = 14
    rng.Interior.Pattern = 2
    rng.Interior.Color = rgbToInt((255, 255, 255))

def create_index_dict(key, periods):
    count = 0
    dct = dict()
    if key == 'По данным 1С':
        pattern1 = ['ДДУ', 'ДДУ', 'ДДУ', 'ДДУ', 'ДДУ', '', 'ДКП', 'ДКП', '', '']
        pattern2 = ['Реализация_ДДУ', 'Реализация_ДДУ', 'ДДУ', 'ДДУ', '', '', 'Реализация_ДКП', 'Реализация_ДКП',
                    'ДКП', 'ДКП']

        for period in periods:
            pattern3 = [period] * 10
            for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
                if count == 0:
                    dct[str(-1)] = ['Index 1', 'Index 2', 'Index 3']
                dct[str(count)] = [ptrn1, ptrn2, ptrn3]
                count += 1
    elif key == 'По данным CRM':
        pattern1 = ['Реализация_ДДУ', 'Реализация_ДДУ', 'ДДУ', 'ДДУ', '', '', 'Реализация_ДКП', 'Реализация_ДКП',
                    'ДКП', 'ДКП']
        for period in periods:
            pattern2 = [period] * 10
            for ptrn1, ptrn2 in zip(pattern1, pattern2):
                if count == 0:
                    dct[str(-1)] = ['Index 1', 'Index 2']
                dct[str(count)] = [ptrn1, ptrn2]
                count += 1
    else:
        for period in periods:
            pattern1 = [period] * 10
            for ptrn1 in pattern1:
                if count == 0:
                    dct[str(-1)] = 'Index 1'
                dct[str(count)] = ptrn1
                count += 1
    return dct

def create_column_dict(key, periods, StartRow, DICT_name):
    column_dict = dict()
    count = 0
    if key == 'По данным 1С':
        pattern1 = []
        pattern2 = ['Реализация по ДДУ', '', '', '', '', '', 'Реализация по ДКП', '', '', '']
        pattern3 = ['Заключение', 'Расторжение', 'Поступление ДС', 'Возврат ДС', 'Прочее движение', '',
                    'Заключение', 'Расторжение', '', '']
        for i in range(len(periods)):
            formula = f'=IF(ISERROR(SEARCH("_",{get_column_letter(5 + 10 * i)}{StartRow + 3},1)),' \
                      f'CONCATENATE({get_column_letter(5 + 10 * i)}{StartRow + 3}," год"),' \
                      f'CONCATENATE(VLOOKUP(MID({get_column_letter(5 + 10 * i)}{StartRow + 3},' \
                      f'1,SEARCH("_",{get_column_letter(5 + 10 * i)}{StartRow + 3})-1)+0,' \
                      f'{DICT_name}!$A:$B,2,0),"-е ",' \
                      f'RIGHT({get_column_letter(5 + 10 * i)}{StartRow + 3},4)," года"))'
            pattern1.append(formula)
            pattern1.extend(['', '', '', '', '', '', '', '', ''])

            for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
                column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
                count += 1
            pattern1 = []

        for _ in periods:
            for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
                column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
                count += 1
    elif key == 'По данным CRM':
        pattern2 = ['Реализация по ДДУ', '', '', '', '', '', 'Реализация по ДКП', '', '', '']
        pattern3 = ['Заключение\n(тыс.руб.)', 'Расторжение\n(тыс.руб.)', 'Заключение\n(кв.м.)', 'Расторжение\n(кв.м.)',
                    '',
                    '', 'Заключение\n(тыс.руб.)', 'Расторжение\n(тыс.руб.)', 'Заключение\n(кв.м.)',
                    'Расторжение\n(кв.м.)']
        for i in range(len(periods)):
            pattern1 = []
            formula = f'=IF(ISERROR(SEARCH("_",{get_column_letter(5 + 10 * i)}{StartRow + 2},1)),' \
                      f'CONCATENATE({get_column_letter(5 + 10 * i)}{StartRow + 2}," год"),' \
                      f'CONCATENATE(VLOOKUP(MID({get_column_letter(5 + 10 * i)}{StartRow + 2},' \
                      f'1,SEARCH("_",{get_column_letter(5 + 10 * i)}{StartRow + 2})-1)+0,' \
                      f'{DICT_name}!$A:$B,2,0),"-е ",' \
                      f'RIGHT({get_column_letter(5 + 10 * i)}{StartRow + 2},4)," года"))'
            pattern1.append(formula)
            pattern1.extend(['', '', '', '', '', '', '', '', ''])
            for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
                column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
                count += 1
    else:
        pattern2 = ['Разница', '', '', '', '', '', '', 'Разница по ДКП', '', '']
        pattern3 = ['Заключение', 'Расторжение', '', '', '', '',
                    'Заключение', 'Расторжение', '', '', ]
        for i in range(len(periods)):
            pattern1 = []
            formula = f'=IF(ISERROR(SEARCH("_",{get_column_letter(5 + 10 * i)}{StartRow + 1},1)),' \
                      f'CONCATENATE({get_column_letter(5 + 10 * i)}{StartRow + 1}," год"),' \
                      f'CONCATENATE(VLOOKUP(MID({get_column_letter(5 + 10 * i)}{StartRow + 1},' \
                      f'1,SEARCH("_",{get_column_letter(5 + 10 * i)}{StartRow + 1})-1)+0,' \
                      f'{DICT_name}!$A:$B,2,0),"-е ",' \
                      f'RIGHT({get_column_letter(5 + 10 * i)}{StartRow + 1},4)," года"))'
            pattern1.append(formula)
            pattern1.extend(['', '', '', '', '', '', '', '', ''])
            for ptrn1, ptrn2, ptrn3 in zip(pattern1, pattern2, pattern3):
                column_dict[str(count)] = [ptrn1, ptrn2, ptrn3]
                count += 1
    return column_dict


def create_value_dict(ws, df, key, prj, periods, StartRow, BIT_row, CRM_row, DDU_name, DKP_name, CRM_name):
    value_dict = dict()
    count = 5
    if key == 'По данным 1С':

        PeriodRow = ws.Range(ws.Cells(StartRow-10, 4), ws.Cells(StartRow, 4)).Find('Index 3').Row
        DDU_FORMULA = f'=SUMIFS({DDU_name}!$AB:$AB,' \
                      f'{DDU_name}!$X:$X,${PeriodRow + 3}:${PeriodRow + 3},' \
                      f'{DDU_name}!$U:$U,${PeriodRow}:${PeriodRow},' \
                      f'{DDU_name}!$V:$V,$C:$C,{DDU_name}!$W:$W,$D:$D)'
        DKP_FORMULA = f'=SUMIFS({DKP_name}!$AB:$AB,' \
                      f'{DKP_name}!$X:$X,${PeriodRow + 3}:${PeriodRow + 3},' \
                      f'{DKP_name}!$U:$U,${PeriodRow}:${PeriodRow},' \
                      f'{DKP_name}!$V:$V,$C:$C,{DKP_name}!$W:$W,$D:$D)'
        for _ in periods:
            for i in range(10):
                if i <= 4:
                    lst = [DDU_FORMULA] * len(df)
                elif i == 5 or i in (8, 9):
                    lst = [''] * len(df)
                else:
                    lst = [DKP_FORMULA] * len(df)
                EdnRow = ws.Range("{0}{1}".format('B', ws.Rows.Count)).End(-4162).Row
                StartSumRow = ws.Range(ws.Cells(StartRow, 2), ws.Cells(EdnRow, 2)).Find(prj).Row - 1
                lst.extend(['', f'=SUM({get_column_letter(count)}{StartSumRow}:{get_column_letter(count)}{StartSumRow + len(df) - 1})'])
                value_dict[str(count)] = lst
                count += 1
    elif key == 'По данным CRM':
        PeriodRow = ws.Range(ws.Cells(StartRow-10, 4), ws.Cells(StartRow, 4)).Find('Index 2').Row
        DDU_FORMULA_METRES_CLOSE = FORMULA_METRES_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                   DOC_TYPE='ДДУ')
        DKP_FORMULA_METRES_CLOSE = FORMULA_METRES_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                   DOC_TYPE='ДКП')
        DDU_FORMULA_METRES_DENIAL = FORMULA_METRES_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                     DOC_TYPE='ДДУ')
        DKP_FORMULA_METRES_DENIAL = FORMULA_METRES_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                     DOC_TYPE='ДКП')
        DDU_FORMULA_MONEY_CLOSE = FORMULA_MONEY_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                 DOC_TYPE='ДДУ')
        DKP_FORMULA_MONEY_CLOSE = FORMULA_MONEY_CLOSE.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                 DOC_TYPE='ДКП')
        DDU_FORMULA_MONEY_DENIAL = FORMULA_MONEY_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                   DOC_TYPE='ДДУ')
        DKP_FORMULA_MONEY_DENIAL = FORMULA_MONEY_DENIAL.substitute(CRM_NAME=CRM_name, SALES_PERIOD=PeriodRow,
                                                                   DOC_TYPE='ДКП')

        FORMULS = [DDU_FORMULA_METRES_CLOSE, DDU_FORMULA_METRES_DENIAL, DDU_FORMULA_MONEY_CLOSE,
                   DDU_FORMULA_MONEY_DENIAL, '', '',
                   DKP_FORMULA_METRES_CLOSE, DKP_FORMULA_METRES_DENIAL, DKP_FORMULA_MONEY_CLOSE,
                   DKP_FORMULA_MONEY_DENIAL]
        for _ in periods:
            for i, formula in enumerate(FORMULS):
                lst = [formula] * len(df)
                EdnRow = ws.Range("{0}{1}".format('B', ws.Rows.Count)).End(-4162).Row
                StartSumRow = ws.Range(ws.Cells(StartRow, 2), ws.Cells(EdnRow, 2)).Find(prj).Row - 1
                lst.extend(['', f'=SUM({get_column_letter(count)}{StartSumRow}:{get_column_letter(count)}{StartSumRow + len(df) - 1})'])
                value_dict[str(count)] = lst
                count += 1
    else:
        for _ in periods:
            for i in range(10):
                if i in (0, 1, 6, 7):
                    lst = [f'={BIT_row + k}:{BIT_row + k}-{CRM_row + k}:{CRM_row + k}' for k in range(len(df))]
                    # value_dict[str(count)] = [f'{BIT_row}:{BIT_row}-{CRM_row}:{CRM_row}'] * 10
                else:
                    lst = [''] * len(df)
                    # value_dict[str(count)] = [''] * 10
                EdnRow = ws.Range("{0}{1}".format('B', ws.Rows.Count)).End(-4162).Row
                StartSumRow = ws.Range(ws.Cells(StartRow, 2), ws.Cells(EdnRow, 2)).Find(prj).Row - 1
                lst.extend(['', f'=SUM({get_column_letter(count)}{StartSumRow}:{get_column_letter(count)}{StartSumRow + len(df) - 1})'])
                value_dict[str(count)] = lst
                count += 1
    return value_dict