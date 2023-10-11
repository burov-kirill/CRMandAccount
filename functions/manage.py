import asyncio
import os
import time
from datetime import datetime
from threading import Thread

from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pandas as pd
import re
from copy import copy
import win32timezone
import string
import win32com.client
from collections import namedtuple
from classes.crm_file import CrmFile
from classes.excel_file import AccountSales, AccountPayment
from functions.create_file import open_and_fill_new_file, fill_data
from settings.logs import log
from settings.templates import PERIOD_FORMULA
import pythoncom


RowsDict = {'По данным 1С': 6, 'По данным CRM': 5, 'Разница': 4,
            'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок': 6,
            'Корректировка кв.м.': 5,
            'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок': 5,
            'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок': 6,
            'Корректировка тыс.руб.': 5,
            'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок': 5,
            'Партнерские продажи, кв. м.': 5, 'Партнерские продажи, тыс. руб.': 5, }

EditFormulasDict = {'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок': 7,
            'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок': 6,
            'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок': 7,
            'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок': 6,
            'Партнерские продажи, кв. м.': 6, 'Партнерские продажи, тыс. руб.': 6,
}

ROMANIAN_NUMBERS = pd.DataFrame({1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI',  7: 'VII',
                    8: 'VIII', 9: 'IX', 10: 'X', 11: 'XI', 12: 'XII'}.items(), columns = ['latin', 'romanian'])

period_dict = {'Месяц': [period for element in [
    [f'1_{year}', f'2_{year}', f'3_{year}', f'4_{year}', f'5_{year}', f'6_{year}', f'7_{year}', f'8_{year}',
     f'9_{year}', f'10_{year}', f'11_{year}', f'12_{year}']
    for year in range(2014, 2020)] for period in element],
               'Квартал': [period for element in [[f'1_{year}', f'2_{year}', f'3_{year}', f'4_{year}']
                                                  for year in range(2014, 2020)] for period in element],
               'Полугодие': [period for element in [[f'1_{year}', f'2_{year}']
                                                    for year in range(2014, 2020)] for period in element],
               'Год': [str(year) for year in range(2014, 2020)]
               }

def add_rows(ws, data_path, prj):
    ws.EnableCalculation = False
    EndRow = get_last_row_from_column(ws, 'B', True)
    new_data = pd.read_excel(data_path)
    for key, value in RowsDict.items():
        ArticleRow = ws.Range(ws.Cells(1, 2), ws.Cells(EndRow, 2)).Find(key).Row
        # Start = ws.Range(ws.Cells(ArticleRow, 2), ws.Cells(EndRow, 2)).Find(prj).Row
        # # Start = get_split_row(ws, EndRow, key) + value
        End = ws.Range(ws.Cells(ArticleRow, 2), ws.Cells(EndRow, 2)).Find('ИТОГО').Row - 1
        EndRow += len(new_data)
        for _ in range(len(new_data)):
            ws.Rows(End).Insert(-4121)
            ws.Rows(End - 1).Copy()
            ws.Rows(End).PasteSpecial(-4123)
            ws.Rows(End).PasteSpecial(-4122)
            End += 1

        ws.Range(ws.Cells(End - len(new_data), 3),  # Cell to start the "paste"
                    ws.Cells(End - len(new_data) + len(new_data.index) - 1,
                                3 + len(new_data.columns) - 1)  # No -1 for the index
                    ).Value = new_data.values
    ws.EnableCalculation = True
def get_excel_range(number):
    if isinstance(number, (int, float)):
        return get_column_letter(number)
    else:
        return column_index_from_string(number)

def get_split_row(ws, EndRow, text = '', StartRow = 1, option = True, column = 'B'):
    split_row = 0
    if option:
        for i in range(StartRow, EndRow+1):
            if ws.Range(f'{column}{i}').Value == text: # 'СВЕРКА 1С и CRM'
                split_row = i
                break
    else:
        for i in range(StartRow, EndRow+1):
            if re.match(r'\d*_?\d{4}\.?\d?', str(ws.Range(f'E{i}').Value)) != None \
                    and str(ws.Range(f'E{i}').Value) == re.match(r'\d*_?\d{4}\.?\d?', str(ws.Range(f'E{i}').Value)).group():
                split_row = i - StartRow
                break
    return split_row

def get_split_col(ws, SearchRow, StartRow = 1, EndRow = 10000):
    value_list = ws.Range(f'{SearchRow}:{SearchRow}').Value[0]
    res_col = 0
    for i in range(StartRow, EndRow):
        if value_list[i] == None:
            res_col = i
            break
    return res_col
def get_last_row_from_column(ws, column='', option = True):
    if option:
        max_end_row = ws.Range("{0}{1}".format(column, ws.Rows.Count)).End(-4162).Row
    else:
        columns = list(string.ascii_uppercase)
        columns.extend(['AA', 'AB'])
        max_end_row = max([ws.Range("{0}{1}".format(col, ws.Rows.Count)).End(-4162).Row for col in columns])
    return max_end_row

def get_last_num(ws):
    EndRow = get_last_row_from_column(ws, 'Q', True)
    last_num = ws.Cells(EndRow, 17).Value
    if isinstance(last_num, (int, float)):
        return last_num
    else:
        return 0
def create_sheet_dict(wb):
    user_data = namedtuple('UserData', ['sheetname', 'object'])
    sheet_dict = dict()
    DDU_check, DKP_check, CRM_check, SUMM_check = False, False, False, False
    for sheet in wb.Sheets:
        if 'ДДУ' in sheet.Name and not DDU_check:
            sheet_dict['AccPay'] = user_data(sheet.Name, AccountPayment)
            DDU_check = True
        elif 'ДКП' in sheet.Name and not DKP_check:
            sheet_dict['AccSales'] = user_data(sheet.Name, AccountSales)
            DKP_check = True
        elif 'CRM' in sheet.Name and not CRM_check:
            CRM_check = True
            sheet_dict['CRM'] = user_data(sheet.Name, CrmFile)
    return sheet_dict

def create_cumsum_column(sheet, df):
    last_num = get_last_num(sheet)
    df['cum_sum'] = df['12_Сумма ДТ'] - df['14_Сумма КТ']
    df['cum_sum'] = df['cum_sum'].cumsum()
    df['cum_sum'] = df['cum_sum'] + last_num
    return df


def create_custom_range(ws, project, range_type, EndRow, SplitRow):
    # project = custom_replace(project)
    custom_range = namedtuple('range', ['type', 'copy', 'past'])
    if range_type in ('AccPay', "AccSales"):
        prj_row = get_split_row(ws, SplitRow, project)
        first_col = get_split_col(ws, prj_row, 1)
        last_col = get_split_col(ws, prj_row, first_col + 1) + 1
        if range_type == 'AccPay':

            copy_range = f'{get_excel_range(first_col)}18:{get_excel_range(first_col)}{SplitRow}'
            past_range = f'{get_excel_range(first_col + 1)}18:{get_excel_range(first_col + 1)}{SplitRow}'
        else:
            copy_range = f'{get_excel_range(last_col-1)}18:{get_excel_range(last_col-1)}{SplitRow}'
            past_range = f'{get_excel_range(last_col)}18' \
                         f':{get_excel_range(last_col)}{SplitRow}'
    else:

        prj_row = get_split_row(ws, EndRow, 'СВЕРКА 1С и CRM', SplitRow + 2)
        prj_row = get_split_row(ws, EndRow, 'Index 3', prj_row, True, "D")
        last_col = get_split_col(ws, prj_row, 5)
        if last_col == 0:
            pass

        copy_range = f'{get_excel_range(last_col-9)}{SplitRow+4}:{get_excel_range(last_col)}{EndRow}'
        past_range = f'{get_excel_range(last_col + 1)}{SplitRow+4}:{get_excel_range(last_col+10)}{EndRow}'

    return custom_range(range_type, copy_range, past_range)

def get_current_periods(ws, rng, table_type, StartCol = 4):
    EndCol = get_split_col(ws, rng, StartCol)
    if table_type == 'AccSales':
        StartCol = EndCol + 1
        EndCol = get_split_col(ws, rng, StartCol)
    raw_values = ws.Range(f'{rng}:{rng}').Value[0][StartCol:EndCol]
    return raw_values
def check_range(ws, rng, user_rng, table_type, StartCol = 4):
    raw_values = get_current_periods(ws, rng, table_type, StartCol)
    if user_rng not in raw_values:
        return True
    else:
        return False
def change_formula(string, column_number):
    if re.search(r'\+([A-Z]+:[A-Z]+)', string) != None:
        string = re.sub(r'\+([A-Z]+:[A-Z]+)', f'+{get_column_letter(column_number-1)}:{get_column_letter(column_number-1)}', string)
    return string

def change_range(sheet,EndRow, user_values, obj, create_new_file):
    for key, value in EditFormulasDict.items():
        AtricleRow = sheet.Range(sheet.Cells(1, 2), sheet.Cells(EndRow, 2)).Find(key).Row
        InitRow = sheet.Range(sheet.Cells(AtricleRow, 2), sheet.Cells(EndRow, 2)).Find(user_values['prj']).Row
        InitCol = get_split_col(sheet, InitRow) + 2
        EndRowNew = sheet.Range(sheet.Cells(InitRow, 2), sheet.Cells(EndRow, 2)).Find('ИТОГО').Row - 2
        EndCol = get_split_col(sheet, InitRow, InitCol)
        formula_frame = pd.DataFrame(sheet.Range(sheet.Cells(InitRow, InitCol), sheet.Cells(EndRowNew, EndCol)).Formula)
        for num_col, col in enumerate(formula_frame.columns, InitCol):
            formula_frame[col] = formula_frame[col].apply(change_formula, args=[num_col])
        sheet.Range(sheet.Cells(InitRow, InitCol), sheet.Cells(InitRow + len(formula_frame.index) - 1, InitCol + len(formula_frame.columns) - 1)).ClearContents()
        sheet.Range(sheet.Cells(InitRow, InitCol), sheet.Cells(InitRow + len(formula_frame.index) - 1, InitCol + len(formula_frame.columns) - 1)).Value = formula_frame.values
def get_subpath(path, i, opt = True):
    while i > 0:
        if opt and path.rfind('\\') == -1:
            path = path[:path.rfind('/')]
        else:
            path = path[:path.rfind('\\')]
        i-=1
    return path

def get_periods(crm_df, ddu_df, dkp_df, option=True, prd = 'Полугодие'):
    crm_period = set(crm_df['Квартал_Год регистрации'])
    ddu_period, dkp_period = set(ddu_df), set(dkp_df)
    crm_period = set(re.findall(r'\d*_?\d{4}', ' '.join(crm_period)))
    ddu_period = set(re.findall(r'\d*_?\d{4}', ' '.join(ddu_period)))
    dkp_period = set(re.findall(r'\d*_?\d{4}', ' '.join(dkp_period)))

    if not option:
        if prd != 'Год':
            return sorted(max([ddu_period, dkp_period, crm_period], key=lambda x: len(x)), key=lambda x: (x.split('_')[1], x.split('_')[0]))
        else:
            return sorted(max([ddu_period, dkp_period, crm_period], key=lambda x: len(x)))

    if len(dkp_period) == len(ddu_period) == len(crm_period):
        return sorted(crm_period, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))
    elif len(crm_period)>len(dkp_period) and len(crm_period)>len(ddu_period):
        recent_periods = list(crm_period.difference(ddu_period).difference(dkp_period))
        return sorted(recent_periods, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))
    else:
        recent_periods = list(ddu_period.difference(crm_period).difference(dkp_period))
        return sorted(recent_periods, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))

def get_queue_frame(crm_file, ddu_file, dkp_file, project, prd):
    project = project.replace('_', ' ')
    ddu_period = ddu_file.additional_frame['Квартал_Год'] if not ddu_file.is_empty  else set()
    dkp_period = dkp_file.additional_frame['Квартал_Год'] if not dkp_file.is_empty  else set()
    crm_df = copy(crm_file.additional_frame)
    crm_df['Проект'] = project
    crm_df = crm_df[['Проект', "Очередь", "Дом"]]
    crm_df['Очередь'] = crm_df['Очередь'].apply(str)
    crm_df['Дом'] = crm_df['Дом'].apply(str)
    result_frame = crm_df.groupby(['Проект', "Очередь", "Дом"], as_index=False).count()
    period = get_periods(crm_file.additional_frame, ddu_period, dkp_period, prd)
    result_frame = result_frame[result_frame['Очередь']!= '']
    return result_frame, period

def extend_list(period_list, current_list, prd, data_list = ''):
    unique_period_list = set(period_list)
    unique_list = set(current_list)
    result_list = unique_period_list.difference(unique_list).union(unique_list)
    if prd != 'Год':
        for year in range(2014, int(max(result_list, key=lambda x: int(x.split('_')[1])).split('_')[1]) + 1):
            if prd == 'Месяц':
                period = range(1, 13)
            elif prd == 'Полугодие':
                period = range(1, 3)
            else:
                period = range(1, 5)
            for p in period:
                if f'{p}_{year}' not in result_list:
                    if data_list != '' and f'{p}_{year}' not in data_list:
                        result_list.add(f'{p}_{year}')
                if data_list != '' and f'{p}_{year}' in data_list and f'{p}_{year}' in result_list:
                    result_list.remove(f'{p}_{year}')
        return sorted(result_list, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))
    else:
        for year in range(2014, int(max(result_list, key=lambda x: int(x)))):
            if f'{year}' not in result_list:
                if data_list != '' and f'{year}' not in data_list:
                    result_list.add(f'{year}')
            if data_list != '' and f'{year}' in data_list and f'{year}' in result_list:
                result_list.remove(f'{year}')
        return sorted(result_list, key=lambda x: int(x))

def get_period(string, period):
    string = str(string)
    if string != '':
        date_string = string.split(' ')[0]
        if re.search(r'\d{4}-\d{2}-\d{2}', date_string) != None:
            pattern = '%Y-%m-%d'
        elif re.search(r'\d{2}\.\d{2}\.\d{4}', date_string) != None:
            pattern = '%d.%m.%Y'
        else:
            return string
        if period == 'Полугодие':
            if datetime.strptime(date_string, pattern).date().month <= 6:
                return '1'
            else:
                return '2'
        elif period == 'Месяц':
            return str(datetime.strptime(date_string, pattern).date().month)
        elif period == 'Квартал':
            return str(pd.Timestamp(datetime.strptime(date_string, pattern).date()).quarter)
        else:
            return str(datetime.strptime(date_string, pattern).date().year)
    else:
        return string

def fill_date_columns(df, period, is_crm = False):
    if is_crm:
        if period != 'Год':
            df['Рег_Период'] = df['registration_date'].apply(get_period, args=[period])
            df['Рег_Год'] = df['registration_date'].apply(get_period, args=['Год'])
            df['Рег_Период_Год'] = df['Рег_Период'] + '_' + df['Рег_Год']
            df['Рас_Период'] = df['cancellation_date'].apply(get_period, args=[period])
            df['Рас_Год'] = df['cancellation_date'].apply(get_period, args=['Год'])
            df['Рас_Период_Год'] = df['Рас_Период'] + '_' + df['Рас_Год']
        else:
            df['Рег_Период'] = ''
            df['Рег_Год'] = df['registration_date'].apply(get_period, args=[period])
            df['Рег_Период_Год'] = df['Рег_Год']
            df['Рас_Период'] = ''
            df['Рас_Год'] = df['cancellation_date'].apply(get_period, args=[period])
            df['Рас_Период_Год'] = df['Рас_Год']
        df = df[['Рег_Период', 'Рег_Год', 'Рег_Период_Год', 'Рас_Период', 'Рас_Год', 'Рас_Период_Год']]
        return df
    else:
        if period != 'Год':
            df['Период'] = df['date'].apply(get_period, args=[period])
            df['Год'] = df['date'].apply(get_period, args=['Год'])
            df['Период_Год'] = df['Период'] + '_' + df['Год']
        else:
            df['Период'] = ''
            df['Год'] = df['date'].apply(get_period, args=[period])
            df['Период_Год'] = df['Год']
        df = df[['Период', 'Год', 'Период_Год']]
        short_df = df['Период_Год']
        return (df, short_df)


def add_columns(sheet, pivot_sheet, user_values, periods, change_period, create_new_file, data_list = ''):
    sheet.EnableCalculation = False
    log.info('Добавление столбцов')
    EndRow = get_last_row_from_column(sheet, 'B', True)
    SplitRow = get_split_row(sheet, EndRow, 'СВЕРКА 1С и CRM') - 2

    DownTableDict = {element: get_split_row(sheet, EndRow, element, SplitRow) + get_split_row(sheet, EndRow,
                                                                                              StartRow=get_split_row(
                                                                                                  sheet, EndRow,
                                                                                                  element, SplitRow),
                                                                                              option=False)
                     for element in ['По данным 1С', 'По данным CRM', 'Разница']}

    UpTableDict = {element: get_split_row(sheet, SplitRow, element) + get_split_row(sheet, SplitRow,
                                                                                    StartRow=get_split_row(sheet,
                                                                                                           SplitRow,
                                                                                                           element),
                                                                                    option=False)
                   for element in
                   ['Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок',
                    'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок']}
    for k, v in pivot_sheet.items():
        if user_values['--CREATE_FILE--'] or change_period:
            if data_list == '':
                if k in ('AccPay', 'AccSales'):
                    data_list = get_current_periods(sheet, UpTableDict[list(UpTableDict.keys())[0]], k)
                else:
                    data_list = get_current_periods(sheet, DownTableDict[list(DownTableDict.keys())[0]], k)
            v = extend_list(periods, v, user_values['--TO_PERIOD--'], data_list)
        for element in v:
            if k in ('AccPay', 'AccSales'):
                check_rng = check_range(sheet, UpTableDict[list(UpTableDict.keys())[0]], element, k)
            else:
                check_rng = check_range(sheet, DownTableDict[list(DownTableDict.keys())[0]], element, k)
            if check_rng:
                obj = create_custom_range(sheet, user_values['prj'], k, EndRow, SplitRow)  # user_values['project']
                if k == 'AccPay':

                    # Thread(target=past_column, args = [sheet, obj.past]).start()
                    # Thread(target=select_cell, args=[sheet]).start()
                    # Thread(target=func2).start()
                    sheet.Range(obj.past).Insert()
                try:
                    sheet.Range(obj.copy).Copy()
                    sheet.Range(obj.past).PasteSpecial(-4123)
                    sheet.Range(obj.past).PasteSpecial(8)

                    sheet.Range(obj.past).PasteSpecial(-4122)
                except Exception as inner_exp:
                    log.exception(f'{inner_exp}\n{k}\n{obj.copy}\n{obj.past}')

                if k in ('AccPay', "AccSales"):
                    for key, value in UpTableDict.items():
                        temp_range = re.sub(r'\d+', str(value), obj.past)
                        sheet.Range(temp_range).Value = element
                else:
                    for key, value in DownTableDict.items():
                        temp_range = re.sub(r'\d+', str(value), obj.past)
                        sheet.Range(temp_range).Value = element

        if k == 'AccPay' and check_rng:
            change_range(sheet, EndRow, user_values, obj, create_new_file or change_period)
    sheet.EnableCalculation = True

def edit_date_string(date_string):
    if date_string != None:
        return str(date_string)
    else:
        return date_string

def create_col_dict(ws, col_lst, add_col_list):
    square_col = get_column_letter(col_lst.index('Площадьобщаяподоговору(бездублей)')+1)
    money_col = get_column_letter(col_lst.index('Цена_дог,руб(бездублей)')+1)
    queue_col = f"{get_column_letter(col_lst.index('Очередь'))}:{get_column_letter(col_lst.index('Очередь'))}"
    house_col = f"{get_column_letter(col_lst.index('КорпусНомер')+1)}:{get_column_letter(col_lst.index('КорпусНомер')+1)}"
    add_square_col = f"{get_column_letter(list(add_col_list).index('Площадь')+2+get_split_col(ws, 4))}"
    add_money_col = f"{get_column_letter(list(add_col_list).index('Сумма') + 2 + get_split_col(ws, 4))}"
    add_queue_col = f"{get_column_letter(list(add_col_list).index('Очередь')+2+get_split_col(ws, 4))}:{get_column_letter(list(add_col_list).index('Очередь')+2+get_split_col(ws, 4))}"
    add_house_col = f"{get_column_letter(list(add_col_list).index('Дом')+2+get_split_col(ws, 4))}:{get_column_letter(list(add_col_list).index('Очередь')+2+get_split_col(ws, 4))}"
    return {'square_col': square_col, 'money_col': money_col, 'queue_col': queue_col,'house_col': house_col,
            'add_queue_col':add_queue_col,'add_house_col':add_house_col, 'add_square_col': add_square_col, 'add_money_col': add_money_col}


def write_new_data(wb, files):
    pivot_sheet = dict()
    for i, obj in enumerate(files):
        if not files[i].is_empty:
            add_col = files[i].additional_column
        if files[i].type_file != 'CRM' and files[i].is_empty == False:
            sheet = wb.Worksheets(files[i].sheet_name)
            StartRow = get_last_row_from_column(sheet, option=False) + 2
            df = files[i].df
            add_df = files[i].additional_frame
            df = create_cumsum_column(sheet, df)
        elif files[i].type_file == 'CRM' and files[i].is_empty == False:
            sheet = wb.Worksheets(files[i].sheet_name)
            if sheet.FilterMode:
                sheet.ShowAllData()
            StartRow = get_last_row_from_column(sheet, option=False) + 2
            sheet.Range(f"B5:BE{StartRow}").ClearContents()
            df = files[i].full_df
            add_df = files[i].additional_frame

            StartRow = 5
        if files[i].is_empty == False:
            if files[i].type_file != 'CRM':
                for col in ['L:L', 'N:N', 'W:W']:
                    sheet.Range(col).NumberFormat = '@'
                    sheet.Range(col).Replace(',', '.')
                for col in ['M:M', 'Q:Q', 'AB:AB', 'O:O']:
                    sheet.Range(col).NumberFormat = '# ##0'
            elif files[i].type_file == 'CRM':
                spl_col = get_split_col(sheet, 4)
                lst = sheet.Range(sheet.Cells(4, 1), sheet.Cells(4, spl_col)).Value
                flat_lst = [item for sublist in lst for item in sublist]
                flat_lst = list(map(lambda x: x.replace(' ', '').replace('\n', ''), flat_lst))
                df = df[flat_lst]
                add_col = spl_col + 2
                col_dict = create_col_dict(sheet, flat_lst, add_df.columns)
                for key, value in col_dict.items():
                    if key not in ('square_col', 'money_col', 'add_square_col', 'add_money_col'):
                        sheet.Range(value).NumberFormat = '@'
                        sheet.Range(value).Replace(',', '.')
                condition = get_column_letter(len(df.columns) + list(add_df.columns).index('Учитывается(нет/да)') + 2)
                add_df['Площадь'] = [f'=IF({condition}{i + 5}="Да",{col_dict["square_col"]}{i + 5},0)' for i in range(len(add_df))]
                add_df['Сумма'] = [f'=IF({condition}{i + 5}="Да",{col_dict["money_col"]}{i + 5},0)/1000' for i in range(len(add_df))]
                add_df['Корректировка м.2'] = [f'=-{col_dict["add_square_col"]}{i+5}' if add_df['Корректировка м.2'][i]!='' else '' for i in range(len(add_df))]
                add_df['Корректировка тыс.руб.'] = [f'=-{col_dict["add_money_col"]}{i + 5}' if add_df['Корректировка м.2'][i] != '' else '' for i in range(len(add_df))]

            sheet.Range(sheet.Cells(StartRow, files[i].column),  # Cell to start the "paste"
                            sheet.Cells(StartRow + len(df.index) - 1,
                                        files[i].column + len(df.columns) - 1)  # No -1 for the index
                            ).Value = df.values

            sheet.Range(sheet.Cells(StartRow, add_col),  # Cell to start the "paste"
                            sheet.Cells(StartRow + len(add_df.index) - 1,
                                        files[i].column + len(df.columns) + len(add_df.columns))  # No -1 for the index
                            ).Value = add_df.values
            UpdateEndRow = get_last_row_from_column(sheet, 'B', True) + 1

            if files[i].type_file != 'CRM':
                sheet.Range(f'B{6}:AB{6}').Copy()
                sheet.Range(f'B{StartRow}:AB{UpdateEndRow}').PasteSpecial(-4122)
            else:
                sheet.Range(sheet.Cells(StartRow, add_col), sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 6)).Interior.Color\
                    = rgbToInt((221, 235, 247))
                sheet.Range(sheet.Cells(StartRow, add_col + 7),sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 8)).Interior.Color\
                    = rgbToInt((255, 255, 0))
                sheet.Range(sheet.Cells(StartRow, add_col + 9),sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 9)).Interior.Color\
                    = rgbToInt((255, 230, 153))
                sheet.Range(sheet.Cells(StartRow, add_col + 10), sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 10)).Interior.Color\
                    = rgbToInt((221, 235, 247))
                sheet.Range(sheet.Cells(StartRow, add_col + 11),sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 11)).Interior.Color\
                    = rgbToInt((255, 230, 153))
                sheet.Range(sheet.Cells(StartRow, add_col + 12), sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 12)).Interior.Color\
                    = rgbToInt((221, 235, 247))
                sheet.Range(sheet.Cells(StartRow, add_col + 16), sheet.Cells(StartRow + len(add_df.index) - 1, add_col + 16)).Interior.Color\
                    = rgbToInt((221, 235, 247))
        if obj.type_file in ('AccPay', 'AccSales'):
            if not obj.is_empty:
                pivot_sheet[obj.type_file] = obj.period
            else:
                pivot_sheet[obj.type_file] = []

    for k, v in pivot_sheet.items():
        if v == []:
            pivot_sheet[k] = max(pivot_sheet.values(), key=lambda x: len(x))

    sheet = wb.Worksheets([ws.Name for ws in wb.Sheets if 'Свод' in ws.Name][0])
    pivot_sheet['DownTable'] = max(pivot_sheet.values(), key=lambda x: len(x))
    return (sheet, pivot_sheet)

def custom_replace(obj):
    if isinstance(obj, list):
        return "".join(obj).replace(" ", "_").replace("-", "_")
    else:
        return obj.replace(" ", "_").replace("-", "_")

def change_data(ws, period, dict_sheet, project):
    lst = ['Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок', 'Корректировка кв.м.',
           'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок',
           'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок','Корректировка тыс.руб.',
           'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок',
           'Партнерские продажи, кв. м.',
           'Партнерские продажи, тыс. руб.', 'По данным 1С', 'По данным CRM', 'Разница']
    StartRow = 1
    EndRow = 1000
    period_row = 1
    data_list = []
    end_col = 1
    for item in lst:
        if item in ('Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок', 'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок', 'По данным 1С', 'По данным CRM', 'Разница'):

            if item == 'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок':
                GlobalEndColumn = get_split_col(ws, ws.Range(f'E{StartRow}:E{EndRow}').Find('1_2014').Row, 5)
            try:
                check = ws.Range(f'E{StartRow}:E{EndRow}').Find('1_2014').Row-ws.Range(f'B{StartRow}:B{EndRow}').Find(item).Row<10
            except Exception as inner_exp:
                log.exception('Обнаружена ссылка на строки периодов')
                check = False
            finally:
                if check:
                    period_row = ws.Range(f'E{StartRow}:E{EndRow}').Find('1_2014').Row
                    StartRow = period_row + 1
                    end_col = get_split_col(ws, period_row, 5)
                else:
                    if item == 'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок':
                        start = ws.Range(f'B:B').Find(item).Row
                        start = ws.Range(f'B{start}:B{EndRow}').Find(project).Row - 3
                        formula_list = [
                            PERIOD_FORMULA.substitute(COLUMN=get_column_letter(i), ROW=period_row, DICT_SHEET=dict_sheet)
                            for i in range(5, end_col + 1)]
                        ws.Range(ws.Cells(start, 5), ws.Cells(start, end_col)).Value = formula_list
                        formula_list = [
                            PERIOD_FORMULA.substitute(COLUMN=get_column_letter(i), ROW=period_row, DICT_SHEET=dict_sheet)
                            for i in range(end_col + 2, end_col + 2 + len(data_list))]
                        ws.Range(ws.Cells(start, end_col + 2),
                                 ws.Cells(start, end_col + 1 + len(data_list))).Value = formula_list
                        continue
                    else:
                        start = ws.Range(f'B:B').Find(item).Row
                        start = ws.Range(f'B{start}:B{EndRow}').Find(project).Row - 3
                if item in ('По данным 1С', 'По данным CRM', 'Разница'):
                    if item == 'По данным 1С':
                        GlobalStartRow = ws.Range(f'B:B').Find('Данные по продажам').Row
                        GlobalStartColumn = get_column_letter(len(data_list) + 5)
                        GlobalEndRow = ws.Range(f'B:B').Find(item).Row
                        if len(data_list)<GlobalEndColumn+1-5:
                            ws.Range(f'{GlobalStartColumn}{GlobalStartRow}:{get_column_letter(GlobalEndColumn)}{GlobalEndRow}').Delete(-4159)
                            GlobalStartColumn = 5 + len(data_list)*2 + 1
                            GlobalEndColumn = 5 + len(data_list)*2 + 1 + GlobalEndColumn - 5-len(data_list)
                            ws.Range(f'{get_column_letter(GlobalStartColumn)}{GlobalStartRow}:{get_column_letter(GlobalEndColumn)}{GlobalEndRow}').Delete(-4159)
                    all_periods = data_list
                    # all_periods = period_dict[period][:(end_col + 1 - 5)//10]
                    periods = []
                    for prd in all_periods:
                        periods.extend([prd]*10)
                    if check:
                        ws.Range(ws.Cells(period_row, 5), ws.Cells(period_row, end_col)).Value = periods
                    formula_list = []
                    column = 5
                    if item == 'Разница':
                        for _ in all_periods:
                            formula_list.append(PERIOD_FORMULA.substitute(COLUMN=get_column_letter(column), ROW=period_row,
                                                                          DICT_SHEET=dict_sheet))
                            column +=6
                            formula_list.extend([''] * 5)
                            formula_list.append(PERIOD_FORMULA.substitute(COLUMN=get_column_letter(column), ROW=period_row,
                                                                          DICT_SHEET=dict_sheet))
                            column += 4
                            formula_list.extend([''] * 3)
                        if check:
                            ws.Range(ws.Cells(period_row + 1, 5), ws.Cells(period_row + 1, end_col)).Value = formula_list
                        else:
                            ws.Range(ws.Cells(start, 5),
                                     ws.Cells(start, end_col)).Value = formula_list
                        if len(formula_list) + 5 < end_col:
                            GlobalStartRow = ws.Range(f'B:B').Find('По данным 1С').Row
                            if check:
                                GlobalEndRow = ws.Range(f'B{period_row}:B1000').Find('ИТОГО').Row
                            else:
                                GlobalEndRow = ws.Range(f'B{start}:B1000').Find('ИТОГО').Row
                            GlobalStartColumn = len(formula_list) + 5
                            GlobalEndColumn = end_col
                            ws.Range(
                                f'{get_column_letter(GlobalStartColumn)}{GlobalStartRow}:{get_column_letter(GlobalEndColumn)}{GlobalEndRow}').Delete(
                                -4159)
                    else:
                        for _ in all_periods:
                            formula_list.append(PERIOD_FORMULA.substitute(COLUMN = get_column_letter(column), ROW = period_row, DICT_SHEET = dict_sheet))
                            formula_list.extend(['']*9)
                            column+=10
                        if check:
                            ws.Range(ws.Cells(period_row + 1, 5), ws.Cells(period_row + 1, end_col)).Value = formula_list
                        else:
                            ws.Range(ws.Cells(start, 5),
                                     ws.Cells(start, end_col)).Value = formula_list


                else:
                    data_list = period_dict[period][:end_col + 1 - 5]
                    ws.Range(ws.Cells(period_row, 5), ws.Cells(period_row, end_col)).Value = data_list
                    ws.Range(ws.Cells(period_row, end_col + 2), ws.Cells(period_row, end_col + 1 + len(data_list))).Value = data_list
                    formula_list = [PERIOD_FORMULA.substitute(COLUMN = get_column_letter(i), ROW = period_row, DICT_SHEET = dict_sheet) for i in range(5, end_col + 1)]
                    ws.Range(ws.Cells(period_row + 2, 5), ws.Cells(period_row + 2, end_col)).Value = formula_list
                    formula_list = [PERIOD_FORMULA.substitute(COLUMN = get_column_letter(i), ROW = period_row, DICT_SHEET = dict_sheet) for i in range(end_col + 2, end_col + 2 + len(data_list))]
                    ws.Range(ws.Cells(period_row + 2, end_col + 2), ws.Cells(period_row + 2, end_col + 1 + len(data_list))).Value = formula_list
        elif item not in ('По данным 1С', 'По данным CRM', 'Разница'):
            start = ws.Range(f'B:B').Find(item).Row
            start=ws.Range(f'B{start}:B{EndRow}').Find(project).Row-3
            formula_list = [PERIOD_FORMULA.substitute(COLUMN=get_column_letter(i), ROW=period_row, DICT_SHEET=dict_sheet)
                            for i in range(5, end_col + 1)]
            ws.Range(ws.Cells(start, 5), ws.Cells(start, end_col)).Value = formula_list
            formula_list = [PERIOD_FORMULA.substitute(COLUMN=get_column_letter(i), ROW=period_row, DICT_SHEET=dict_sheet)
                for i in range(end_col + 2, end_col + 2 + len(data_list))]
            ws.Range(ws.Cells(start, end_col + 2), ws.Cells(start, end_col + 1 + len(data_list))).Value = formula_list
    return data_list

def rgbToInt(rgb):
    if isinstance(rgb, tuple):
        colorInt = rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)
        return colorInt
    else:
        return rgb
def main_func(user_values):
    create_new_file = user_values['--CREATE_FILE--']
    periods = []
    check_report = True
    files = []
    pivot_sheet = dict()
    # out.Update('Считывание файлов')
    # progress_value = queue.get_nowait()
    # pg_bar.UpdateBar(progress_value)
    # pg_bar.UpdateBar(1)
    if user_values['--CREATE_FILE--']:
        log.info(f'Считывание данных для создания итогового файла')
        crm_file = CrmFile(user_values['CRM'], 'CRM', f'CRM_{user_values["prj"].replace("-", "_").replace(" ", "_")}',False if user_values['CRM']!='' else True, user_values['spt'],user_values['--TO_PERIOD--'])
        ddu_file = AccountPayment(user_values['AccPay'], 'AccPay', f'ДДУ_{user_values["prj"].replace("-", "_").replace(" ", "_")}', False if user_values['AccPay']!='' else True,
                                    user_values['spt'], user_values['--TO_PERIOD--'])
        dkp_file = AccountSales(user_values['AccSales'], 'AccSales', f'ДКП_{user_values["prj"].replace("-", "_").replace(" ", "_")}', False if user_values['AccSales']!='' else True,
                                  user_values['spt'], user_values['--TO_PERIOD--'])
        files = [ddu_file, dkp_file, crm_file]
        path = user_values['save_folder']
        name = f'/СверкаCRM_{user_values["prj"]}.xlsb'
        user_values['SummaryFile'] = os.path.abspath((path+name))
        df, periods = get_queue_frame(crm_file, ddu_file, dkp_file, user_values['prj'], user_values["--TO_PERIOD--"])
        open_and_fill_new_file(path, name, user_values['prj'], df, user_values["--TO_PERIOD--"])
        user_values['SummaryFile'] = os.path.abspath((path+name))
    Excel = win32com.client.Dispatch("Excel.Application", pythoncom.CoInitialize())
    Excel.DisplayAlerts = False
    Excel.Visible = False
    Excel.ScreenUpdating = False
    Excel.EnableEvents = False
    Excel.AskToUpdateLinks = False
    wb = Excel.Workbooks.Open(user_values['SummaryFile'])
    change_period = False
    if user_values['--FROM_PERIOD--'] != user_values['--TO_PERIOD--'] and not user_values['--CREATE_FILE--']:
        log.info(f'Изменение исходных данных в случае выбора разных временных периодов')
        change_period = True
        if not os.path.isdir(get_subpath(user_values['SummaryFile'],1)+'/Итоговые_файлы'):
            os.mkdir(get_subpath(user_values['SummaryFile'],1)+'/Итоговые_файлы')
        new_file_name = get_subpath(user_values['SummaryFile'],1)+f'/Итоговые_файлы/СверкаCRM_{"".join(user_values["prj"])}_{user_values["--TO_PERIOD--"]}.xlsb'
        wb.SaveCopyAs(os.path.abspath((new_file_name)))
        wb.Close()
        wb = Excel.Workbooks.Open(os.path.abspath((new_file_name)))
        user_values['SummaryFile'] = os.path.abspath((new_file_name))
        DDU_check, DKP_check, CRM_check, SUMM_check = False, False, False, False
        for sheet in wb.Sheets:
            if 'ДДУ' in sheet.Name or 'ДКП' in sheet.Name:
                EndRow = get_last_row_from_column(sheet, 'B', True) + 1
                ws = wb.Worksheets(sheet.Name)
                data = ws.Range(ws.Cells(6, 2), ws.Cells(EndRow, 2))
                period_data = pd.DataFrame(data.Value, columns=['date'])
                period_data.fillna('', inplace=True)

                period_data, short_df = fill_date_columns(period_data, user_values['--TO_PERIOD--'])
                if 'ДДУ' in sheet.Name and not DDU_check:
                    DDU_check = True
                    sheet.Name = sheet.Name.replace(" ", "_").replace("-", "_")
                    DDU_name = sheet.Name
                    ddu_period_df = short_df
                elif 'ДКП' in sheet.Name and not DKP_check:
                    DKP_check = True
                    sheet.Name = sheet.Name.replace(" ", "_").replace("-", "_")
                    DKP_name = sheet.Name
                    dkp_period_df = short_df
                ws.Range(ws.Cells(6, 19),  # Cell to start the "paste"
                            ws.Cells(6 + len(period_data.index) - 1,
                                        19 + len(period_data.columns) - 1)  # No -1 for the index
                            ).Value = period_data.values
            elif 'Свод' in sheet.Name and not SUMM_check:
                SUMM_check = True
                sheet.Name = sheet.Name.replace(" ", "_").replace("-", "_")
                del_name = sheet.Name
            elif 'CRM' in sheet.Name and not CRM_check:
                CRM_check = True
                sheet.Name = sheet.Name.replace(" ", "_").replace("-", "_")
                CRM_name = sheet.Name
                ws = wb.Worksheets(sheet.Name)
                if ws.FilterMode:
                    ws.ShowAllData()
                EndRow = get_last_row_from_column(ws, 'B', True)
                data = ws.Range(ws.Cells(5, 39), ws.Cells(EndRow, 39))
                crm_period_df = pd.DataFrame(columns=['Квартал_Год регистрации'])
                crm_period_df.fillna('', inplace=True)
                if user_values['--REVIEW--']:
                    reg_col = ws.Range('4:4').Find('Дата_рег договора').Column
                    cancel_col = ws.Range('4:4').Find('Дата раст.-я').Column
                    data = ws.Range(ws.Cells(5, reg_col), ws.Cells(EndRow, cancel_col))
                    data = [(edit_date_string(element[0]), edit_date_string(element[1]))  for element in data.Value]
                    period_data = pd.DataFrame(data, columns=['registration_date', 'cancellation_date'])
                    period_data.fillna('', inplace=True)
                    period_data = fill_date_columns(period_data, user_values['--TO_PERIOD--'], True)
                    spl_col = get_split_col(ws, 4) + 2
                    ws.Range(ws.Cells(5, spl_col),  # Cell to start the "paste"
                             ws.Cells(5 + len(period_data.index) - 1,
                                      spl_col + len(period_data.columns) - 1)  # No -1 for the index
                             ).Value = period_data.values

        ws = wb.Worksheets(del_name)
        BeginRow = get_split_row(ws, get_last_row_from_column(ws, 'B', True), user_values['prj'])
        EndRow = get_split_row(ws, get_last_row_from_column(ws, 'B', True), 'ИТОГО', BeginRow) - 2
        prj_data = ws.Range(ws.Cells(BeginRow, 2), ws.Cells(EndRow, 4)).Value
        df = pd.DataFrame(prj_data, columns=['Проект', 'Очередь', 'Дом'])
        # wb.Worksheets(del_name).Delete()
        periods = get_periods(crm_period_df, ddu_period_df, set() if dkp_period_df.empty else dkp_period_df, False, user_values['--TO_PERIOD--'])
        # period = periods[0]
        # periods = periods[1:]
        add = wb.Sheets.Add(Before=None, After=wb.Sheets(wb.Sheets.count))
        # add.Name = 'Свод_' + custom_replace(user_values["prj"])
        if f'Словарь_{custom_replace(user_values["prj"])}' not in [sheet.Name for sheet in wb.Sheets]:
            add = wb.Sheets.Add(Before=None, After=wb.Sheets(wb.Sheets.count))
            add.Name = f'Словарь_{custom_replace(user_values["prj"])}'
            DICT_NAME = f'Словарь_{custom_replace(user_values["prj"])}'
            ws = wb.Worksheets(f'Словарь_{custom_replace(user_values["prj"])}')
            ws.Range(ws.Cells(1, 1),  # Cell to start the "paste"
                     ws.Cells(1 + len(ROMANIAN_NUMBERS.index) - 1,
                                 1 + len(ROMANIAN_NUMBERS.columns) - 1)  # No -1 for the index
                     ).Value = ROMANIAN_NUMBERS.values
        else:
            DICT_NAME = f'Словарь_{custom_replace(user_values["prj"])}'
        # ws = wb.Worksheets('Свод_' +custom_replace(user_values["prj"]))
        ws = wb.Worksheets(del_name)
        ws.Activate()
        Excel.ActiveWindow.DisplayGridlines = False
        prj_for_formula = user_values['prj'].replace('-', '_').replace(' ', '_')
        prj_for_data = user_values['prj']
        data_list = change_data(ws, user_values["--TO_PERIOD--"], DICT_NAME, prj_for_data)
        # fill_data(ws, df, DDU_name, DKP_name, CRM_name, DICT_NAME , prj_for_formula, user_values['--TO_PERIOD--'], prj_for_data)
        if user_values['--REVIEW--']:
            log.info(f'Оформление ревью')
            # queue.put(2)
            # out.Update('Считывание файлов')
            # progress_value = queue.get_nowait()
            # pg_bar.UpdateBar(progress_value)
            # pg_bar.UpdateBar(2)
            if user_values['--TO_PERIOD--'] != 'Год':
                pivot_sheet = {
                                'AccPay': [min(periods, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))],
                                'AccSales': [min(periods, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))],
                                'DownTable': [min(periods, key=lambda x: (int(x.split('_')[1]), int(x.split('_')[0])))]
                                }
            else:
                pivot_sheet = {
                                'AccPay': [min(periods, key=lambda x: int(x))],
                                'AccSales': [min(periods, key=lambda x: int(x))],
                                'DownTable':[min(periods, key=lambda x: int(x))]
                                }
            add_columns(ws, pivot_sheet, user_values, periods, change_period, create_new_file, data_list)
            wb.Save()
            wb.Close()
            Excel.Quit()
            return user_values

    if not user_values['--CREATE_FILE--']:
        sheet_dict = create_sheet_dict(wb)
        for k, v in sheet_dict.items():
            if user_values[k] == '':
                is_empty = True
            else:
                is_empty = False
            files.append(v.object(user_values[k], k, v.sheetname, is_empty, user_values['spt'], user_values['--TO_PERIOD--']))

    # out.Update('Запись в файл')
    log.info(f'Запись в файл')
    # progress_value = queue.get_nowait()
    # pg_bar.UpdateBar(progress_value)
    # pg_bar.UpdateBar(3)
    sheet, pivot_sheet = write_new_data(wb, files)

    # for i, obj in enumerate(files):
    #     add_col = files[i].additional_column
    #     if files[i].type_file != 'CRM' and files[i].is_empty == False:
    #         sheet = wb.Worksheets(files[i].sheet_name)
    #         StartRow = get_last_row_from_column(sheet, option=False) + 2
    #         df = files[i].df
    #         add_df = files[i].additional_frame
    #         if files[i].type_file == 'AccPay':
    #             df = create_cumsum_column(sheet, df)
    #     elif files[i].type_file == 'CRM' and files[i].is_empty == False:
    #         sheet = wb.Worksheets(files[i].sheet_name)
    #         if sheet.FilterMode:
    #             sheet.ShowAllData()
    #         StartRow = get_last_row_from_column(sheet, option=False) + 2
    #         sheet.Range(f"B5:BE{StartRow}").ClearContents()
    #         df = files[i].full_df
    #         add_df = files[i].additional_frame
    #         StartRow = 5
    #     if files[i].is_empty == False:
    #         if files[i].type_file != 'CRM':
    #             for col in ['L:L', 'N:N', 'W:W']:
    #                 sheet.Range(col).NumberFormat = '@'
    #                 sheet.Range(col).Replace(',', '.')
    #             for col in ['M:M', 'Q:Q', 'AB:AB', 'O:O']:
    #                 sheet.Range(col).NumberFormat = '# ##0'
    #         else:
    #             spl_col = get_split_col(sheet, 4)
    #             lst = sheet.Range(sheet.Cells(4, 1), sheet.Cells(4, spl_col)).Value
    #             flat_lst = [item for sublist in lst for item in sublist]
    #             flat_lst = list(map(lambda x: x.replace(' ', '').replace('\n', ''), flat_lst))
    #             df = df[flat_lst]
    #             add_col = spl_col + 2
    #
    #             for col in ['T:T', 'U:U', 'AR:AR', 'AS:AS']:
    #                 sheet.Range(col).NumberFormat = '@'
    #                 sheet.Range(col).Replace(',', '.')
    #
    #
    #         sheet.Range(sheet.Cells(StartRow, files[i].column),  # Cell to start the "paste"
    #                     sheet.Cells(StartRow + len(df.index) - 1,
    #                                 files[i].column + len(df.columns) - 1)  # No -1 for the index
    #                     ).Value = df.values
    #
    #         sheet.Range(sheet.Cells(StartRow, add_col),  # Cell to start the "paste"
    #                     sheet.Cells(StartRow + len(add_df.index) - 1,
    #                                 add_col + len(add_df.columns) - 1)  # No -1 for the index
    #                     ).Value = add_df.values
    #         UpdateEndRow = get_last_row_from_column(sheet, 'B', True) + 1
    #         if files[i].type_file != 'CRM':
    #             sheet.Range(f'B{6}:AB{6}').Copy()
    #             sheet.Range(f'B{StartRow}:AB{UpdateEndRow}').PasteSpecial(-4122)
    #
    #
    #     if obj.type_file in ('AccPay', 'AccSales'):
    #         if not obj.is_empty:
    #             pivot_sheet[obj.type_file] = obj.period
    #         else:
    #             pivot_sheet[obj.type_file] = []
    #
    # for k, v in pivot_sheet.items():
    #     if v == []:
    #         pivot_sheet[k] = max(pivot_sheet.values(), key=lambda x: len(x))
    #
    # sheet = wb.Worksheets([ws.Name for ws in wb.Sheets if 'Свод' in ws.Name][0])
    # pivot_sheet['DownTable'] = max(pivot_sheet.values(), key=lambda x: len(x))

    add_columns(sheet, pivot_sheet, user_values, periods, change_period, create_new_file)
    # out.Update('Оформление СВОДа: добавление столбцов')
    # progress_value = queue.get_nowait()
    # pg_bar.UpdateBar(progress_value)
    # pg_bar.UpdateBar(4)


    if user_values['--ADD_STRING--'] and  'new_data' in user_values.keys() and user_values['new_data'] != '':
        add_rows(sheet, user_values['new_data'], user_values['prj'])
    wb.Save()
    wb.Close()
    Excel.DisplayAlerts = True
    Excel.Visible = True
    Excel.ScreenUpdating = True
    Excel.EnableEvents = True
    Excel.AskToUpdateLinks = True
    Excel.Quit()
    return user_values
