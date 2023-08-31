import os
import sys
import time
from collections import namedtuple

from openpyxl.utils.cell import get_column_letter
import pandas as pd
import win32com.client
from win32com.client import constants as c


period = '1_2014'
PRJ = 'АЛХИМОВО'
df = pd.DataFrame({'Проект': ['АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО' , 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО' , 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО' , 'АЛХИМОВО', 'АЛХИМОВО', 'АЛХИМОВО'], 'Очередь': [1, 1, 2, 2, 2, 2, 3, 3, 4, 4, 4, 4, 4, 4, 5, 5],
                   'Дом': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]})


ADD_COLUMNS = ['Полугодие','Год','Квартал_Год','Очередь','Дом','Тип','Контрагент','Тип контрагента','Договор','Сумма']
BIT_COLUMNS = ['Дата', 'Операция', 'Операция_1', 'Аналитика_Дт', 'Аналитика_Дт', 'Аналитика_Дт', 'Аналитика_Кт',
               'Аналитика_Кт', 'Аналитика_Кт', 'Аналитика_Кт', 'Счет_Дт', 'Сумма_Дт', 'Счет_Кт', 'Сумма_Кт', 'Тип операции', 'Сальдо накопительным итогом']
ADD_CRM_COLUMNS = ['Квартал регистрации договора', 'Год регистрации', 'Квартал_Год регистрации',
                            'Квартал расторжения договора', 'Год расторжения', 'Квартал_Год расторжения',
                            'Тип договора',
                            'Очередь', 'Дом', 'Учитывается(нет/да)', 'Контрагент', 'Тип контрагента (Партнер или нет)',
                            'Договор', 'Площадь', 'Сумма', 'Комментарий', 'Расторгнут?(да/нет)', 'Корректировка м.2',
                            'Корректировка тыс.руб.']

CRM_COLUMNS = ['Договор, #', '№ Договора', 'Тип договора', 'Дата заключения', 'Дата_рег договора',
                      'Дата раст.-я',
                      'Дата АПП', 'Дата ПАПП', 'Контрагент', 'Цена_дог,руб (дубли)',
                      'Площадь общая по договору (дубли)', 'Цена_дог,руб(без дублей)',
                      'Площадь общая по договору (без дублей)', 'Адрес дома', 'Застройка', 'Очередь', 'Помещение Тип',
                      'Помещение Под Тип',
                      'Помещение Студ', 'Корпус Номер', 'Этаж', '№кв.', 'Площадь общая по обмерам БТИ', '# пом.',
                      'Помещение', 'Правообладатель Название',
                      'Прав.Тип', '## плт. граф.', '### плт. факт', 'Дата_плт.', 'Сумма_плт., руб (без дублей)',
                      'График Платежей Задолженность (без дублей)',
                      'График Платежей Дней Просрочки', 'Дата платежа факт', 'Сумма, руб']

ROMANIAN_NUMBERS = pd.DataFrame({1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI',  7: 'VII',
                    8: 'VIII', 9: 'IX', 10: 'X', 11: 'XI', 12: 'XII' }.items(), columns = ['latin', 'romanian'])
d = {
    'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок': ((112, 48, 160), (255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Корректировка кв.м.': ((112, 48, 160), (255, 255, 255), 'Корректировка продаж по ДДУ'),
    'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок': ((112, 48, 160), (255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок': ((82, 170, 49),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Корректировка тыс.руб.': ((82, 170, 49),(255, 255, 255), 'Корректировка продаж по ДДУ'),
    'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок': ((82, 170, 49),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Партнерские продажи, кв. м.': ((4, 65, 149),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
    'Партнерские продажи, тыс. руб.': ((4, 65, 149),(255, 255, 255), 'Заключено ДДУ (накопительным итогом)'),
}

period_dict = {'Месяц': [period for element in [[f'1_{year}', f'2_{year}', f'3_{year}', f'4_{year}', f'5_{year}', f'6_{year}', f'7_{year}', f'8_{year}', f'9_{year}', f'10_{year}', f'11_{year}', f'12_{year}']
                         for year in range(2014, 2020)] for period in element],
               'Квартал': [period for element in [[f'1_{year}', f'2_{year}', f'3_{year}', f'4_{year}']
                         for year in range(2014, 2020)] for period in element],
               'Полугодие': [period for element in [[f'1_{year}', f'2_{year}']
                         for year in range(2014, 2020)] for period in element],
               'Год': [str(year) for year in range(2014, 2020)]
               }

def create_bottom_dict(user_period):
    bottom_dict = {
        'По данным 1С': ((112, 48, 160), (255, 255, 255), period_dict[user_period], {
            'Реализация по ДДУ': ['Заключение', 'Расторжение', 'Поступление ДС', 'Возврат ДС', 'Прочее \nдвижение'],
            'Реализация по ДКП': ['Заключение', 'Расторжение']},
                         {
                             'Index 1': ['ДДУ', 'ДКП'],
                             'Index 2': ['Реализация_ДДУ', 'Реализация_ДКП'],
                             'Index 3': period_dict[user_period]
                         }),
        'По данным CRM': ((82, 170, 49), (255, 255, 255), period_dict[user_period], {
            'Реализация по ДДУ': ["Заключение \n(тыс.руб.)", "Расторжение \n(тыс.руб.)", "Заключение \n(кв.м.)",
                                  "Расторжение \n(кв.м.)"],
            'Реализация по ДКП': ["Заключение \n(тыс.руб.)", "Расторжение \n(тыс.руб.)", "Заключение \n(кв.м.)",
                                  "Расторжение \n(кв.м.)"]},
                          {
                              'Index 1': ['Реализация_ДДУ', 'Реализация_ДКП'],
                              'Index 2': period_dict[user_period],
                          }),
        'Разница': ((4, 65, 149), (255, 255, 255), period_dict[user_period], {'Разница': ['Заключение', 'Расторжение'],
                                                            'Разница по ДКП': ['Заключение', 'Расторжение']},
                    {
                        'Index 1': period_dict[user_period]
                    }),
    }
    return bottom_dict
def create_excel_file(path, name):
    Excel = win32com.client.Dispatch("Excel.Application")
    Excel.DisplayAlerts = False
    Excel.Visible = False
    try:
        wb = Excel.Workbooks.Add()
        path = os.path.abspath((path+name))
        wb.SaveAs(f'{path}', 50)
    except Exception as exp:
        Excel.Quit()
        sys.exit()
    else:
        wb.Save()
        wb.Close()
        Excel.Quit()
        return path

def open_and_fill_new_file(path, name, prj, df, period):
    filename = create_excel_file(path, name)
    Excel = win32com.client.Dispatch("Excel.Application")
    Excel.DisplayAlerts = False
    Excel.Visible = False
    if Excel.ReferenceStyle == 1:
        Excel.ReferenceStyle = 0
    wb = Excel.Workbooks.Open(filename)

    # sheet = wb.Worksheets('Лист1')
    # sheet.Cells(2.1).Value = "win32com"
    prj = prj.replace('-', '_')
    DDU_sheet, DKP_sheet, CRM_sheet, RES_sheet = create_sheet(wb, prj)
    ws = wb.Worksheets(f'Словарь_{prj}')
    ws.Range(ws.Cells(1, 1),  # Cell to start the "paste"
             ws.Cells(1 + len(ROMANIAN_NUMBERS.index) - 1,
                      1 + len(ROMANIAN_NUMBERS.columns) - 1)  # No -1 for the index
             ).Value = ROMANIAN_NUMBERS.values
    sheet_data = namedtuple('SheetData', ['columns', 'add_columns', 'head', 'init_row',
                                          'init_col', 'interior_color', 'font_color', 'width', 'height'])
    wb.Worksheets(1).Delete()
    sheet_dict = {
        DDU_sheet: sheet_data(BIT_COLUMNS, ADD_COLUMNS, 'Карточка счета 76.33 (1С)', 5, 2, (112, 48, 160), (255, 255, 255), 10, 45),
        DKP_sheet: sheet_data(BIT_COLUMNS, ADD_COLUMNS, 'Карточка счета 90.01.1 (1С)', 5, 2, (112, 48, 160), (255, 255, 255), 10, 60),
        CRM_sheet: sheet_data(CRM_COLUMNS, ADD_CRM_COLUMNS, '', 4, 1, (255, 255, 255), (0, 0, 0), 13, 90)
    }
    create_columns(Excel, wb,  sheet_dict)
    ws = wb.Worksheets(RES_sheet)
    ws.Activate()
    Excel.ActiveWindow.DisplayGridlines = False
    ws.Cells.Font.Name = "Arial"
    ws.Cells.Font.Size = 10
    fill_data(ws, df, DDU_sheet, DKP_sheet, CRM_sheet, prj, period)
    ws = wb.Worksheets(RES_sheet)
    ws.Activate()
    wb.Worksheets(RES_sheet).Tab.Color = rgbToInt((128, 128, 128))
    wb.Worksheets(DDU_sheet).Tab.Color = rgbToInt((4, 65, 149))
    wb.Worksheets(DKP_sheet).Tab.Color = rgbToInt((112, 48, 160))
    wb.Worksheets(CRM_sheet).Tab.Color = rgbToInt((82, 170, 49))
    wb.Save()
    wb.Close()
    Excel.Quit()


def create_sheet(wb, prj):
    # удалить старый лист
    # передавать еще список имен или генерировать его
    sheetnames = ['Свод', 'ДДУ', "ДКП", "CRM", 'Словарь']
    for name in sheetnames:
        add = wb.Sheets.Add(Before=None, After=wb.Sheets(wb.Sheets.count))
        if name == 'Словарь':
            add.Name = name + '_' + prj.replace('-', '_')
        else:
            add.Name = name + '_' + prj
    return (f'ДДУ_{prj}', f"ДКП_{prj}", f"CRM_{prj}", f'СВОД_{prj}')


def create_columns(Excel, wb, sheet_dict):

    for key, value in sheet_dict.items():

        ws = wb.Worksheets(key)
        ws.Activate()
        Excel.ActiveWindow.DisplayGridlines = False
        ws.Rows(value.init_row).RowHeight = value.height
        for i, element in enumerate(value.columns):
            ws.Columns(value.init_col+i).ColumnWidth = value.width
            ws.Cells(value.init_row, value.init_col+i).Value = element
            ws.Cells(value.init_row, value.init_col+i).WrapText = True
            align_cells(ws.Cells(value.init_row, value.init_col + i), value.interior_color, value.font_color, False)
        if 'CRM' not in key:
            rng = ws.Range(ws.Cells(value.init_row - 1, value.init_col), ws.Cells(value.init_row, value.init_col+i))
        else:
            rng = ws.Range(ws.Cells(value.init_row, value.init_col), ws.Cells(value.init_row, value.init_col + i))
        for border_id in range(7, 13):
            if border_id in (11, 12) and 'CRM' not in key:
                rng.Borders(border_id).LineStyle = -4119
                rng.Borders(border_id).Weight = 1
            else:
                rng.Borders(border_id).LineStyle = 1
                rng.Borders(border_id).Weight = 2
        ws.Cells(value.init_row - 1, value.init_col).Value = value.head
        align_cells(ws.Range(ws.Cells(value.init_row - 1, value.init_col), ws.Cells(value.init_row - 1, value.init_col+i)),
                    value.interior_color, value.font_color)
        for j, element in enumerate(value.add_columns, i + 2):
            ws.Columns(value.init_col + j).ColumnWidth = value.width
            ws.Cells(value.init_row, value.init_col + j).Value = element
            ws.Cells(value.init_row, value.init_col + j).WrapText = True
            align_cells(ws.Cells(value.init_row, value.init_col + j), rgbToInt((82, 170, 49)), rgbToInt((255, 255, 255)), False)

        rng = ws.Range(ws.Cells(value.init_row, value.init_col+i+2), ws.Cells(value.init_row, value.init_col+j))
        for border_id in range(7, 13):
            if border_id in (11, 12):
                rng.Borders(border_id).LineStyle = -4119
                rng.Borders(border_id).Weight = 1
            else:
                rng.Borders(border_id).LineStyle = 1
                rng.Borders(border_id).Weight = 2

def fill_data(ws,df, DDU_name, DKP_name, CRM_name, prj, period):
    bottom_dict = create_bottom_dict(period)
    periods = period_dict[period]
    ws.Cells(19,2).Value = "Данные по продажам"
    decorate_cells(ws.Range(ws.Cells(19,2), ws.Cells(19,4+len(periods))), rgbToInt((57, 115, 179)), rgbToInt((255, 255, 255)), 30, 'Arial', True)
    decorate_cells(ws.Range(ws.Cells(19, 4+len(periods)+ 2),ws.Cells(19, 4+len(periods)*2+ 1)), rgbToInt((57, 115, 179)), rgbToInt((255, 255, 255)), 30,
                   'Arial', True)

    StartRow = 47 + len(d) + 2*len(d) + 4*len(d) + len(d)*len(df)

    ws.Cells(StartRow, 2).Value = "СВЕРКА 1С и CRM"
    ws.Cells(StartRow, 2).Interior.Color = rgbToInt((255, 255, 255))
    ws.Cells(StartRow, 2).Font.Color = rgbToInt((0, 0, 0))
    ws.Cells(StartRow, 2).Font.Size = 30
    ws.Cells(StartRow, 2).Font.Name = 'Arial'
    ws.Cells(StartRow, 2).Font.Bold = True
    ws.Cells(StartRow, 2).Font.Italic = True
    StartRow+=2
    subhead_font_color = (0, 0, 0)
    subhead_interior_color = (255, 255, 255)
    BIT_row, CRM_row = 0, 0
    for k, v in bottom_dict.items():
        ws.Cells(StartRow, 2).Value = k
        decorate_cells(ws.Range(ws.Cells(StartRow, 2), ws.Cells(StartRow, 5)), rgbToInt(subhead_interior_color),
                       rgbToInt(subhead_font_color), 20, 'Arial', True)
        if k == 'По данным 1С':
            BIT_row = StartRow + 7
        elif k == 'По данным CRM':
            CRM_row = StartRow + 6
        for key, value in v[4].items():
            StartRow+=1
            ws.Cells(StartRow, 4).Value = key
            align_cells(ws.Cells(StartRow, 4), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
            length = len(v[4][f'Index {len(v[4].keys())}'])
            if (key == 'Index 1' and k == 'Разница') or (key == 'Index 2' and k == 'По данным CRM') or (key == 'Index 3' and k == 'По данным 1С'):
                col = 0
                for period_elem in value:
                    for i in range(10):
                        ws.Cells(StartRow, 5+i + col).Value = period_elem
                        align_cells(ws.Cells(StartRow, 5 + i + col), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                    False)
                    col+=10
            # if len(value) == 1:
            #     for i in range(10):
            #         ws.Cells(StartRow, 5+i).Value = value[0]
            #         align_cells(ws.Cells(StartRow, 5 + i), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
            #                     False)
            else:
                for col in range(length):
                    for i in range(10):
                        if k == 'По данным 1С' and key == 'Index 1':
                            if i in range(5):
                                ws.Cells(StartRow, 5 + i + col*10).Value = value[0]
                                align_cells(ws.Cells(StartRow, 5 + i + col*10), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                            False)
                            elif i in (6, 7):
                                ws.Cells(StartRow, 5 + i + col*10).Value = value[1]
                                align_cells(ws.Cells(StartRow, 5 + i + col*10), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                            False)
                            else:
                                continue
                        else:
                            if i in (0, 1):
                                ws.Cells(StartRow, 5 + i + col*10).Value = value[0]
                                align_cells(ws.Cells(StartRow, 5 + i + col*10), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                            False)
                            elif i in (2, 3):
                                ws.Cells(StartRow, 5 + i + col*10).Value = 'ДДУ'
                                align_cells(ws.Cells(StartRow, 5 + i + col*10), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                            False)
                            elif i in (4, 5):
                                continue
                            elif i in (6, 7):
                                ws.Cells(StartRow, 5 + i + col*10).Value = value[1]
                                align_cells(ws.Cells(StartRow, 5 + i + col*10), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                            False)
                            else:
                                ws.Cells(StartRow, 5 + i + col*10).Value = 'ДКП'
                                align_cells(ws.Cells(StartRow, 5 + i + col*10), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)),
                                            False)

        StartRow += 1

        # цикл на индексы
        for i, subhead in enumerate(['Компания', 'Очередь', 'Дом', v[2]], 2):
            if i == 5:
                count = 0
                for p, prd in enumerate(v[2], i):
                    formula = '=ЕСЛИ(ЕОШИБКА(ПОИСК("_";IA227;1));СЦЕП(IA227;" год");СЦЕП(ВПР(ПСТР(IA227;1;ПОИСК("_";IA227;1)-1)+0;Тех_Список!A:B;2;0);"-е ";ПРАВСИМВ(IA227;4);" года"))'
                    ws.Cells(StartRow, p+count*10-int(bool(count))*count).Value = f'=IF(ISERROR(SEARCH("_",{get_column_letter(p+count*10-int(bool(count))*count)}{StartRow - 1},1)),' \
                                                  f'CONCATENATE({get_column_letter(p+count*10-int(bool(count))*count)}{StartRow - 1}," год"),CONCATENATE(VLOOKUP(MID({get_column_letter(p+count*10-int(bool(count))*count)}{StartRow - 1},' \
                                                  f'1,SEARCH("_",{get_column_letter(p+count*10-int(bool(count))*count)}{StartRow - 1})-1)+0,Словарь_{prj.replace("-","_")}!$A:$B,2,0),"-е ",' \
                                                  f'RIGHT({get_column_letter(p+count*10-int(bool(count))*count)}{StartRow - 1},4)," года"))'
                    # ws.Cells(StartRow,
                    #          i).Value = f'=CONCAT(IF(LEFT({get_column_letter(i)}{StartRow - 1},1)="1","I-е ","II-е "),RIGHT({get_column_letter(i)}{StartRow - 1},4)," года")'
                    # ws.Cells(StartRow,
                    #          i).Value = f'=CONCAT(IF(LEFT({get_column_letter(i)}{StartRow - 1},1)="1","I-е ","II-е "),RIGHT({get_column_letter(i)}{StartRow - 1},4)," года")'
                    if k !='Разница':
                        align_cells(ws.Range(ws.Cells(StartRow, p+count*10-int(bool(count))*count), ws.Cells(StartRow, p+count*10-int(bool(count))*count+9)), v[0], v[1])
                    else:
                        align_cells(ws.Range(ws.Cells(StartRow, p+count*10-int(bool(count))*count), ws.Cells(StartRow, p + 1 + count*10-int(bool(count))*count)), v[0], v[1])
                        ws.Cells(StartRow,
                                 p +count*10 + 6-int(bool(count))*count).Value = f'=IF(ISERROR(SEARCH("_",{get_column_letter(p+6+count*10-int(bool(count))*count)}{StartRow - 1},1)),' \
                                            f'CONCATENATE({get_column_letter(p+6+count*10-int(bool(count))*count)}{StartRow - 1}," год"),CONCATENATE(VLOOKUP(MID({get_column_letter(p+6+count*10-int(bool(count))*count)}{StartRow - 1},' \
                                            f'1,SEARCH("_",{get_column_letter(p+6+count*10-int(bool(count))*count)}{StartRow - 1})-1)+0,Словарь_{prj.replace("-","_")}!$A:$B,2,0),"-е ",RIGHT({get_column_letter(p+6+count*10-int(bool(count))*count)}{StartRow - 1},4)," года"))'
                        # ws.Cells(StartRow,
                        #          i + 6).Value = f'=CONCAT(IF(LEFT({get_column_letter(i+6)}{StartRow - 1},1)="1","I-е ","II-е "),RIGHT({get_column_letter(i+6)}{StartRow - 1},4)," года")'
                        align_cells(ws.Range(ws.Cells(StartRow, p +count*10 + 6-int(bool(count))*count), ws.Cells(StartRow, p +count*10 + 7-int(bool(count))*count)), v[0], v[1])
                    # align_cells(ws.Cells(StartRow, i), v[0], v[1])
                    col = p + count*10-int(bool(count)*count)
                    for key, value in v[3].items():
                        ws.Cells(StartRow + 1, col).Value = key
                        if k != 'Разница':
                            align_cells(ws.Range(ws.Cells(StartRow+1, col),ws.Cells(StartRow+1, col + select_count_col(key))), v[0], v[1])
                        else:
                            align_cells(
                                ws.Range(ws.Cells(StartRow + 1, col), ws.Cells(StartRow + 1, col + 1)),
                                v[0], v[1])
                        for j, head in enumerate(value, col):
                            ws.Cells(StartRow + 2, j).Value = head
                            align_cells(ws.Cells(StartRow + 2, j), v[0], v[1])
                            for row in range(len(df)):
                                ws.Cells(StartRow + 3 + row, j).NumberFormat = '# ###;(# ###);"-"'
                                if k == 'По данным 1С':
                                    if 'ДДУ' in key:
                                        ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({DDU_name}!$AB:$AB,' \
                                                                                f'{DDU_name}!$X:$X,${StartRow + 2}:${StartRow + 2},' \
                                                                                f'{DDU_name}!$U:$U,${StartRow-1}:${StartRow-1},' \
                                                                                f'{DDU_name}!$V:$V,$C:$C,{DDU_name}!$W:$W,$D:$D)'
                                    else:
                                        ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({DKP_name}!$AB:$AB,' \
                                                                                f'{DKP_name}!$X:$X,${StartRow + 2}:${StartRow + 2},' \
                                                                                f'{DKP_name}!$U:$U,${StartRow-1}:${StartRow-1},' \
                                                                                f'{DKP_name}!$V:$V,$C:$C,{DKP_name}!$W:$W,$D:$D)'
                                elif k == 'По данным CRM':
                                    if 'ДДУ' in key:
                                        if 'тыс.руб.' in head:
                                            if 'Заключение' in head:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AY:$AY,' \
                                                                                        f'{CRM_name}!$AM:$AM,${StartRow-1}:${StartRow-1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДДУ")'
                                            else:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AY:$AY,' \
                                                                                        f'{CRM_name}!$AP:$AP,${StartRow - 1}:${StartRow - 1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДДУ")'
                                        else:
                                            if 'Заключение' in head:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AX:$AX,' \
                                                                                        f'{CRM_name}!$AM:$AM,${StartRow - 1}:${StartRow - 1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДДУ")'
                                            else:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AX:$AX,' \
                                                                                        f'{CRM_name}!$AP:$AP,${StartRow - 1}:${StartRow - 1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДДУ")'
                                    else:
                                        if 'тыс.руб.' in head:
                                            if 'Заключение' in head:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AY:$AY,' \
                                                                                        f'{CRM_name}!$AM:$AM,${StartRow-1}:${StartRow-1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДКП")'
                                            else:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AY:$AY,' \
                                                                                        f'{CRM_name}!$AP:$AP,${StartRow - 1}:${StartRow - 1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДКП")'
                                        else:
                                            if 'Заключение' in head:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AX:$AX,' \
                                                                                        f'{CRM_name}!$AM:$AM,${StartRow - 1}:${StartRow - 1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДКП")'
                                            else:
                                                ws.Cells(StartRow + 3 + row, j).Formula = f'=SUMIFS({CRM_name}!$AX:$AX,' \
                                                                                        f'{CRM_name}!$AP:$AP,${StartRow - 1}:${StartRow - 1},{CRM_name}' \
                                                                                        f'!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,' \
                                                                                        f'{CRM_name}!$AQ:$AQ,"ДКП")'

                                else:
                                    # ws.Cells(StartRow + 3 + row, j).Value = f'=@134:134-@147:147'
                                    ws.Cells(StartRow + 3 + row, j).Value = f'={BIT_row+row}:{BIT_row+row}-{CRM_row+row}:{CRM_row+row}'

                            ws.Cells(StartRow + 3 + len(df), j).Value = ''  # format
                            ws.Cells(StartRow + 3 + len(df) + 1,
                                     j).Value = f'=SUM({get_column_letter(j)}{StartRow+3}:{get_column_letter(j)}{StartRow + 3 + len(df)})'
                            ws.Cells(StartRow + 3 + len(df) + 1,
                                     j).Font.Bold = True

                        if head == 'Прочее \nдвижение':
                            col = j + 2
                        elif head == 'Расторжение \n(кв.м.)':
                            col = j + 3
                        elif head == 'Расторжение' and key == 'Разница':
                            col = j + 5

                    count+=1
            else:
                ws.Cells(StartRow, i).Value = subhead
                MergeRange = ws.Range(ws.Cells(StartRow, i), ws.Cells(StartRow + 2, i))
                align_cells(MergeRange, v[0], v[1])


        StartRow+=3
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(df.index) - 1,
                          2 + len(df.columns) - 1)  # No -1 for the index
                 ).NumberFormat = '@'
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                    ws.Cells(StartRow + len(df.index) - 1,
                                2 + len(df.columns) - 1)  # No -1 for the index
                    ).Value = df.values
        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(df.index) - 1,
                          2 + len(df.columns) - 1)  # No -1 for the index
                 ).NumberFormat = '@'

        ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                 ws.Cells(StartRow + len(df.index) - 1,
                          2 + len(df.columns) - 1)  # No -1 for the index
                 ).Replace(',', '.')
        rng = ws.Range(ws.Cells(StartRow-3, 2), (ws.Cells(StartRow+len(df) + 1, 10*length+4)))

        set_border(ws, rng, StartRow - 3, 2, StartRow + len(df)+1, 10*length+4, rgbToInt(v[0]), True)
        if k == 'По данным 1С':
            for i in range(length):
                rng = ws.Range(ws.Cells(StartRow - 1, 10+10*i), ws.Cells(StartRow+len(df) - 1, 10+10*i))
                right_rng = ws.Range(ws.Cells(StartRow - 1, 13+10*i), ws.Cells(StartRow + len(df) - 1, 14+10*i))
                right_rng.Interior.Pattern = 14
                right_rng.Interior.PatternThemeColor = 2
                rng.Interior.Pattern = 14
                rng.Interior.PatternThemeColor = 2
        elif k == 'По данным CRM':
            for i in range(length):
                rng = ws.Range(ws.Cells(StartRow - 1, 9+10*i), ws.Cells(StartRow + len(df) - 1, 10+10*i))
                rng.Interior.Pattern = 14
                rng.Interior.PatternThemeColor = 2
        else:
            for i in range(length):
                rng = ws.Range(ws.Cells(StartRow - 3, 7+10*i), ws.Cells(StartRow + len(df) - 1, 10+10*i))
                right_rng = ws.Range(ws.Cells(StartRow - 3, 13+10*i), ws.Cells(StartRow + len(df) - 1, 14+10*i))
                right_rng.Interior.Pattern = 14
                right_rng.Interior.PatternThemeColor = 2
                rng.Interior.Pattern = 14
                rng.Interior.PatternThemeColor = 2

        ws.Cells(StartRow+len(df), 2).Value = 'Если необходимо вставить дополнительные очереди/дома, то добавьте строку перед текущей'
        ws.Cells(StartRow + len(df), 2).Font.Color = rgbToInt((255, 0, 0))
        ws.Cells(StartRow + len(df), 2).Font.Bold = True

        ws.Cells(StartRow + len(df) + 1,
                 2).Value = 'ИТОГО'
        ws.Cells(StartRow + len(df) + 1,
                 2).Font.Bold = True


        StartRow+=len(df) + 3

    StartRow = 21
    for k, v in d.items():

        ws.Cells(StartRow, 2).Value = k
        decorate_cells(ws.Range(ws.Cells(StartRow, 2), ws.Cells(StartRow, 4 + len(periods))), rgbToInt(v[0]), rgbToInt(v[1]), 20,
                           'Arial', True)
        decorate_cells(ws.Range(ws.Cells(StartRow, 4 + len(periods) + 2), ws.Cells(StartRow, 4 + len(periods)*2 + 1)), rgbToInt(v[0]), rgbToInt(v[1]), 20,
                       'Arial', True)
        StartRow += 1
        for prd, period in enumerate(periods):
            if k in ('Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок',
                         'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок'):
                if k == 'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок':
                    Sales_metres_row = StartRow + 1
                else:
                    Sales_rubles_row = StartRow + 1
                ws.Cells(StartRow + 1,
                         5+prd).Value = period
                ws.Cells(StartRow + 1,
                         5+len(periods)+ 1 +prd).Value = period
                # ws.Cells(StartRow + 1, 5).Value = f'=CONCAT(IF(LEFT({Sales_metres_row},1)="1","I-е ","II-е "),RIGHT({Sales_metres_row},4)," года")'
                # ws.Cells(StartRow + 1, 7).Value = f'=CONCAT(IF(LEFT({Sales_metres_row},1)="1","I-е ","II-е "),RIGHT({Sales_metres_row},4)," года")'
                align_cells(ws.Cells(StartRow + 1, 5+prd), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
                align_cells(ws.Cells(StartRow + 1, 5+len(periods)+ prd + 1), rgbToInt((255, 255, 255)), rgbToInt((191, 191, 191)), False)
                # ws.Cells(StartRow + 1, 5).Font.Color = rgbToInt((191, 191, 191))
            elif k == 'Корректировка кв.м.':
                Sales_corrective_row = StartRow+6
            elif k == 'Корректировка тыс.руб.':
                Sales_corrective_rubles_row = StartRow+6
            StartRow += 2

            for i, subhead in enumerate(['Компания', 'Очередь', 'Дом', v[2]], 2):
                if subhead == v[2]:
                    header = v[2]
                    # ws.Cells(StartRow, i+2).Value = subhead
                    # align_cells(ws.Cells(StartRow, i + 2), v[0], v[1])
                    #
                    # align_cells(ws.Cells(StartRow, i), v[0], v[1])
                    ws.Cells(StartRow + 1, i + prd).Value = f'=IF(ISERROR(SEARCH("_",{get_column_letter(i + prd)}{Sales_metres_row},1)),' \
                                                  f'CONCATENATE({get_column_letter(i + prd)}{Sales_metres_row}," год"),CONCATENATE(VLOOKUP(MID({get_column_letter(i + prd)}{Sales_metres_row},' \
                                                  f'1,SEARCH("_",{get_column_letter(i + prd)}{Sales_metres_row})-1)+0,Словарь_{prj.replace("-","_")}!$A:$B,2,0),"-е ",' \
                                                  f'RIGHT({get_column_letter(i + prd)}{Sales_metres_row},4)," года"))'

                    ws.Cells(StartRow + 1, i + len(periods) + prd + 1).Value = f'=IF(ISERROR(SEARCH("_",{get_column_letter(i + len(periods) + prd + 1)}{Sales_metres_row},1)),' \
                                                  f'CONCATENATE({get_column_letter(i + len(periods) + prd + 1)}{Sales_metres_row}," год"),CONCATENATE(VLOOKUP(MID({get_column_letter(i + len(periods) + prd + 1)}{Sales_metres_row},' \
                                                  f'1,SEARCH("_",{get_column_letter(i + len(periods) + prd + 1)}{Sales_metres_row})-1)+0,Словарь_{prj.replace("-","_")}!$A:$B,2,0),"-е ",' \
                                                  f'RIGHT({get_column_letter(i + len(periods) + prd + 1)}{Sales_metres_row},4)," года"))'
                    # ws.Cells(StartRow + 1, i).Value = f'=CONCAT(IF(LEFT({get_column_letter(i)}{Sales_metres_row},1)="1","I-е ","II-е "),RIGHT({get_column_letter(i)}{Sales_metres_row},4)," года")'
                    # ws.Cells(StartRow + 1, i+2).Value = f'=CONCAT(IF(LEFT({get_column_letter(i+2)}{Sales_metres_row},1)="1","I-е ","II-е "),RIGHT({get_column_letter(i+2)}{Sales_metres_row},4)," года")'
                    MergeRange = ws.Range(ws.Cells(StartRow + 1, i + prd), ws.Cells(StartRow + 3, i + prd))
                    RightMergeRange = ws.Range(ws.Cells(StartRow + 1, i + len(periods) + prd + 1), ws.Cells(StartRow + 3, i + len(periods) + prd + 1))
                    align_cells(RightMergeRange, v[0], v[1])
                else:
                    ws.Cells(StartRow, i).Value = subhead
                    MergeRange = ws.Range(ws.Cells(StartRow, i), ws.Cells(StartRow + 3, i))
                align_cells(MergeRange, v[0], v[1])

            StartRow += 4
            ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                     ws.Cells(StartRow + len(df.index) - 1,
                              2 + len(df.columns) - 1)  # No -1 for the index
                     ).NumberFormat = '@'
            ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                         ws.Cells(StartRow + len(df.index) - 1,
                                  2 + len(df.columns) - 1)  # No -1 for the index
                         ).Value = df.values
            ws.Range(ws.Cells(StartRow, 2),  # Cell to start the "paste"
                     ws.Cells(StartRow + len(df.index) - 1,
                              2 + len(df.columns) - 1)  # No -1 for the index
                     ).Replace(',', '.')
            for i in range(len(df)):
                ws.Cells(StartRow + i, 5 + prd).Interior.Color = rgbToInt((221, 235, 247))
                ws.Cells(StartRow + i, 5 + prd + len(periods) + 1).Interior.Color = rgbToInt((221, 235, 247))
                ws.Cells(StartRow + i, 5 + prd).NumberFormat = '# ###;(# ###);"-"'
                ws.Cells(StartRow + i, 5 + prd + len(periods) + 1).NumberFormat = '# ###;(# ###);"-"'

                if k == 'Продажи кв.м. (накопительный итог) без учета ВГО и дополнительных корректировок':
                    ws.Cells(StartRow + i, 5 + prd).Formula = f'=IF({get_column_letter(5 + prd - 1)}${StartRow-4}="Дом",SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДДУ"),' \
                                                      f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДДУ")+{get_column_letter(5 + prd - 1)}:{get_column_letter(5 + prd - 1)})'

                    ws.Cells(StartRow + i, 5 + prd + len(periods) + 1).Formula = f'=IF({get_column_letter(5 + prd + len(periods))}{StartRow + i}="",SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДКП"),' \
                                                      f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДКП")+{get_column_letter(5 + prd + len(periods))}:{get_column_letter(5 + prd + len(periods))})'
                elif k == 'Корректировка кв.м.':
                    ws.Cells(StartRow + i,
                             5 + prd).Formula = f'=SUMIFS({CRM_name}!$BB:$BB,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AQ:$AQ,"ДДУ")'

                    ws.Cells(StartRow + i,
                             5 + prd + len(periods) + 1).Formula = f'=SUMIFS({CRM_name}!$BB:$BB,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AQ:$AQ,"ДКП")'
                elif k == 'Продажи кв.м. (накопительный итог) с учетом ВГО и дополнительных корректировок':
                    ws.Cells(StartRow+i, 5 + prd).Value = f'=IF({get_column_letter(5 + prd - 1)}${StartRow-4}="Дом", SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДДУ")+{Sales_corrective_row + i}:{Sales_corrective_row + i},' \
                                                    f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДДУ")+{Sales_corrective_row + i}:{Sales_corrective_row + i}+{get_column_letter(5 + prd - 1)}:{get_column_letter(5 + prd -1)})'

                    ws.Cells(StartRow+i, 5 + prd + len(periods) + 1).Formula = f'=IF({get_column_letter(5 + prd + len(periods))}{StartRow + i}="", SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДКП")+{Sales_corrective_row + i}:{Sales_corrective_row + i},' \
                                                    f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"ДКП")+{Sales_corrective_row + i}:{Sales_corrective_row + i}+{get_column_letter(5 + prd + len(periods))}:{get_column_letter(5 + prd + len(periods))})'
                elif k == 'Продажи тыс. руб. (накопительный итог) без учета ВГО и дополнительных корректировок':
                    ws.Cells(StartRow + i, 5 + prd).Formula = f'=IF({get_column_letter(5 + prd - 1)}${StartRow-4}="Дом", SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"Реализация_ДДУ"),' \
                                                      f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"Реализация_ДДУ")+{get_column_letter(5 + prd - 1)}:{get_column_letter(5 + prd - 1)})'

                    ws.Cells(StartRow + i, 5 + prd + len(periods) + 1).Formula = f'=IF({get_column_letter(5 + prd + len(periods))}{StartRow + i}="", SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"Реализация_ДКП"),' \
                                                      f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"Реализация_ДКП")+{get_column_letter(5 + prd + len(periods))}:{get_column_letter(5 + prd + len(periods))})'
                elif k == 'Корректировка тыс.руб.':
                    ws.Cells(StartRow + i, 5 + prd).Formula = f'=SUMIFS({CRM_name}!$BC:$BC,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AQ:$AQ,"ДДУ")'
                    ws.Cells(StartRow + i,
                             5 + prd + len(periods) + 1).Formula = f'=SUMIFS({CRM_name}!$BC:$BC,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AQ:$AQ,"ДКП")'
                elif k == 'Продажи тыс. руб. (накопительный итог) с учетом ВГО и дополнительных корректировок':
                    ws.Cells(StartRow + i, 5 + prd).Formula = f'=IF({get_column_letter(5 + prd - 1)}${StartRow-4}="Дом", SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"Реализация_ДДУ")+{Sales_corrective_rubles_row + i}:{Sales_corrective_rubles_row + i},' \
                                                      f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row-4}:${CRM_row-4},${Sales_metres_row}:${Sales_metres_row},${CRM_row-5}:${CRM_row-5},"Реализация_ДДУ")+{Sales_corrective_rubles_row + i}:{Sales_corrective_rubles_row + i}+{get_column_letter(5 + prd - 1)}:{get_column_letter(5 + prd - 1)})'

                    ws.Cells(StartRow + i,
                             5 + prd + len(periods) + 1).Formula = f'=IF({get_column_letter(5 + prd + len(periods))}{StartRow + i}="", SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row - 4}:${CRM_row - 4},${Sales_metres_row}:${Sales_metres_row},${CRM_row - 5}:${CRM_row - 5},"Реализация_ДКП")+{Sales_corrective_rubles_row + i}:{Sales_corrective_rubles_row + i},' \
                                        f'SUMIFS({CRM_row + i}:{CRM_row + i},${CRM_row - 4}:${CRM_row - 4},${Sales_metres_row}:${Sales_metres_row},${CRM_row - 5}:${CRM_row - 5},"Реализация_ДКП")+{Sales_corrective_rubles_row + i}:{Sales_corrective_rubles_row + i}+{get_column_letter(5 + prd + len(periods))}:{get_column_letter(5 + prd + len(periods))})'

                elif k == 'Партнерские продажи, кв. м.':
                    ws.Cells(StartRow + i, 5 + prd).Formula = f'=IF({get_column_letter(5 + prd - 1)}${StartRow-4}="Дом", SUMIFS({CRM_name}!$AX:$AX,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ")+' \
                                                      f'SUMIFS({CRM_name}!$BB:$BB,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ"),' \
                                                      f'SUMIFS({CRM_name}!$AX:$AX,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ")+' \
                                                      f'SUMIFS({CRM_name}!$BB:$BB,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ")+{get_column_letter(5 + prd - 1)}:{get_column_letter(5 + prd - 1)})'

                    ws.Cells(StartRow + i,
                             5 + prd + len(periods) + 1).Formula = f'=IF({get_column_letter(5 + prd + len(periods))}{StartRow + i}="", SUMIFS({CRM_name}!$AX:$AX,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП")+' \
                                        f'SUMIFS({CRM_name}!$BB:$BB,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП"),' \
                                        f'SUMIFS({CRM_name}!$AX:$AX,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП")+' \
                                        f'SUMIFS({CRM_name}!$BB:$BB,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП")+{get_column_letter(5 + prd + len(periods))}:{get_column_letter(5 + prd + len(periods))})'
                elif k == 'Партнерские продажи, тыс. руб.':
                    ws.Cells(StartRow + i, 5 + prd).Formula = f'=IF({get_column_letter(5 + prd - 1)}${StartRow-4}="Дом", SUMIFS({CRM_name}!$AY:$AY,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ")+' \
                                                      f'SUMIFS({CRM_name}!$BC:$BC,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ"),' \
                                                      f'SUMIFS({CRM_name}!$AY:$AY,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ")+' \
                                                      f'SUMIFS({CRM_name}!$BC:$BC,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДДУ")+{get_column_letter(5 + prd - 1)}:{get_column_letter(5 + prd - 1)})'

                    ws.Cells(StartRow + i,
                             5 + prd + len(periods) + 1).Formula = f'=IF({get_column_letter(5 + prd + len(periods))}{StartRow + i}="", SUMIFS({CRM_name}!$AY:$AY,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП")+' \
                                        f'SUMIFS({CRM_name}!$BC:$BC,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП"),' \
                                        f'SUMIFS({CRM_name}!$AY:$AY,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП")+' \
                                        f'SUMIFS({CRM_name}!$BC:$BC,{CRM_name}!$AT:$AT,"Да",{CRM_name}!$AR:$AR,$C:$C,{CRM_name}!$AS:$AS,$D:$D,{CRM_name}!$AM:$AM,${Sales_metres_row}:${Sales_metres_row},{CRM_name}!$AV:$AV,"Да",{CRM_name}!$AQ:$AQ,"ДКП")+{get_column_letter(5 + prd + len(periods))}:{get_column_letter(5 + prd + len(periods))})'
            StartRow -= 6
        StartRow+=6
        ws.Cells(StartRow + len(df) + 1, 2).Value = 'ИТОГО'
        ws.Cells(StartRow + len(df) + 1, 2).Font.Bold = True
        ws.Cells(StartRow-4, 5).Value = header
        rng = ws.Range(ws.Cells(StartRow - 4, 5),ws.Cells(StartRow-4, 5 + len(periods) - 1))
        rng.HorizontalAlignment = 7
        decorate_cells(rng, v[0], v[1], 10, 'Arial', False)
        # ws.Cells(StartRow-4, 5).Value = header
        # align_cells(ws.Range(ws.Cells(StartRow, 5),ws.Cells(StartRow, 4 + len(periods))), v[0], v[1])
        ws.Cells(StartRow-4, 5 + len(periods) + 1).Value = header
        rng = ws.Range(ws.Cells(StartRow-4, 5 + len(periods) + 1),ws.Cells(StartRow-4, 5 + len(periods)*2))
        rng.HorizontalAlignment = 7
        decorate_cells(rng, v[0], v[1], 10, 'Arial', False)
        # align_cells(ws.Cells(StartRow + len(df), 5 + i), v[0], v[1])

        for i in range(len(periods)):
            ws.Cells(StartRow + len(df), 5 + i).Value = '' # format

            ws.Cells(StartRow + len(df)+1, 5 + i).Value = f'=SUM({get_column_letter(5 + i)}{StartRow}:{get_column_letter(5 + i)}{StartRow + len(df)})'
            ws.Cells(StartRow + len(df) + 1, 5 + i).Font.Bold = True

            ws.Cells(StartRow + len(df), 5 + i + len(periods) + 1).Value = ''  # format
            ws.Cells(StartRow + len(df) + 1, 5 + i + len(periods) + 1).Value = f'=SUM({get_column_letter(5 + i + len(periods) + 1)}{StartRow}:{get_column_letter(5 + i + len(periods) + 1)}{StartRow + len(df)})'
            ws.Cells(StartRow + len(df) + 1, 5 + i + len(periods) + 1).Font.Bold = True

        rng = ws.Range(ws.Cells(StartRow - 4, 2), (ws.Cells(StartRow + len(df)+1, 4 + len(periods))))
        right_rng = ws.Range(ws.Cells(StartRow - 4, 4 + len(periods) + 2), (ws.Cells(StartRow + len(df)+1, 4 + len(periods)*2 + 1)))
        set_border(ws, rng, StartRow - 4, 2, StartRow + len(df)+1, 4 + len(periods), rgbToInt(v[0]), True)
        # set_border(ws, rng, StartRow - 4, 2, StartRow + len(df) + 1, 5, True)
        set_border(ws, right_rng, StartRow - 4, 4 + len(periods) + 2, StartRow + len(df)+1, 4+len(periods)*2 + 1, rgbToInt(v[0]), True)


        StartRow += len(df) + 3
        # for col in ['B', 'C', 'D', 'E', 'G']:
        #     ws.Range(f'{col}:{col}').EntireColumn.AutoFit()
    # for col in ['C:C', 'D:D']:
    #     ws.Range(col).NumberFormat = '@'
    #     ws.Range(col).Replace(',', '.')
    ws.Cells(2, 2).Value = f'СЗ САМОЛЕТ - {prj}'
    decorate_cells(ws.Cells(2, 2), (255, 255, 255), (0, 0, 0), 20, 'Arial', True, True)
    ws.Cells(4, 2).Value = 'Проверка полноты отражения данных'
    decorate_cells(ws.Cells(4, 2), (255, 255, 255), (0, 0, 0), 15, 'Arial', True, False, True)
    ws.Cells(6, 2).Value = 'Сумма продаж, тыс.руб.'
    decorate_cells(ws.Cells(6, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', True)
    ws.Cells(8, 2).Value = 'ПРОДАЖИ'
    rng = ws.Range(ws.Cells(8, 2), ws.Cells(9, 2))
    align_cells(rng, (112, 48, 160), (255, 255, 255))
    ws.Cells(10, 2).Value = 'Сумма продаж по ДДУ'
    decorate_cells(ws.Cells(10, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', False)
    ws.Cells(11, 2).Value = 'Сумма продаж по ДКП'
    decorate_cells(ws.Cells(11, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', False)
    ws.Cells(12, 2).Value = 'ИТОГО'
    decorate_cells(ws.Cells(12, 2), (255, 255, 255), (0, 0, 0), 10, 'Arial', False)

    ws.Cells(8, 3).Value = 'СУММА ПО'
    rng = ws.Range(ws.Cells(8, 3),  ws.Cells(8, 5))
    align_cells(rng, (82, 170, 49),(255, 255, 255))
    ws.Cells(9, 3).Value = '1С'
    rng = ws.Range(ws.Cells(9, 3), ws.Cells(9, 5))
    align_cells(rng, (82, 170, 49),(255, 255, 255),  False)
    # decorate_cells(ws.Cells(9, 3), (82, 170, 49),(255, 255, 255), 10, 'Arial', True)
    ws.Cells(10,
             3).Value = f'=SUMIFS({DDU_name}!AB:AB,{DDU_name}!X:X,"Заключение")+SUMIFS({DDU_name}!AB:AB,{DDU_name}!X:X,"Расторжение")'
    ws.Cells(10, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11, 3).Value = f'=SUMIFS({DKP_name}!AB:AB,{DKP_name}!X:X,"Заключение")'
    ws.Cells(11, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(12, 3).Value = f'=SUM(C10:C11)'
    ws.Cells(12, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(9, 4).Value = 'CRM'
    # decorate_cells(ws.Cells(9, 4), (82, 170, 49), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10,
             4).Value = f'=SUMIFS({CRM_name}!AY:AY,{CRM_name}!AQ:AQ,"ДДУ")-SUMIFS({CRM_name}!AY:AY,{CRM_name}!BA:BA,"Да")'
    ws.Cells(10, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11, 4).Value = f'=SUMIFS({CRM_name}!AY:AY,{CRM_name}!AQ:AQ,"ДКП")'
    ws.Cells(11, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(12, 4).Value = f'=SUM(D10:D11)'
    ws.Cells(12, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(9, 5).Value = 'СВОД'
    # decorate_cells(ws.Cells(9, 5), (82, 170, 49), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 5).Value = f'=SUMIFS({BIT_row+len(df)+1}:{BIT_row+len(df)+1},{BIT_row-5}:{BIT_row-5},"Реализация_ДДУ")'
    ws.Cells(10, 5).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11,
             5).Value = f'=SUMIFS({BIT_row + len(df) + 1}:{BIT_row + len(df) + 1},{BIT_row - 5}:{BIT_row - 5},"Реализация_ДКП")'
    ws.Cells(11, 5).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(12, 5).Value = f'=SUM(E10:E11)'
    ws.Cells(12, 5).NumberFormat = '# ###;(# ###);"-"'

    ws.Cells(8, 6).Value = 'ПРОВЕРКА'
    rng = ws.Range(ws.Cells(8, 6), ws.Cells(8, 8))
    align_cells(rng, (4, 65, 149),(255, 255, 255))
    ws.Cells(9, 6).Value = '1С VS CRM'
    rng = ws.Range(ws.Cells(9, 6), ws.Cells(9, 8))
    align_cells(rng, (4, 65, 149),(255, 255, 255),  False)
    # decorate_cells(ws.Cells(9, 6), (4, 65, 149),(255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 6).Value = '=IF(ROUNDDOWN(C10,-3)=ROUNDDOWN(D10,-3),"КОРРЕКТНО","ОШИБКА")'

    ws.Cells(11, 6).Value = '=IF(ROUND(C11,1)=ROUND(D11,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(9, 7).Value = '1С VS СВОД'
    # decorate_cells(ws.Cells(9, 7), (4, 65, 149), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 7).Value = '=IF(ROUND(C10,1)=ROUND(E10,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(11, 7).Value = '=IF(ROUND(C11,1)=ROUND(E11,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(9, 8).Value = 'CRM VS СВОД'
    # decorate_cells(ws.Cells(9, 8), (4, 65, 149), (255, 255, 255), 10, 'Arial', True)
    ws.Cells(10, 8).Value = '=IF(ROUNDDOWN(D10,-3)=ROUNDDOWN(E10,-3),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(11, 8).Value = '=IF(ROUND(D11,1)=ROUND(E11,1),"КОРРЕКТНО","ОШИБКА")'
    rng = ws.Range(ws.Cells(10, 6), ws.Cells(11, 8))
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108

    ws.Cells(11, 9).Value = '=C11-D11'
    ws.Cells(11, 9).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(11, 10).Value = 'n/s'
    ws.Cells(10, 9).Value = '=C10-D10'
    ws.Cells(10, 9).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(10, 10).Value = 'n/s'
    rng = ws.Range(ws.Cells(10, 9), ws.Cells(11, 10))
    decorate_cells(rng, (255, 255, 255), (255, 0, 0), 10, 'Arial', True, False, True)

    ws.Cells(14, 3).Value = '1С'
    ws.Cells(14, 4).Value = 'СВОД'
    ws.Cells(14, 5).Value = 'ПРОВЕРКА'
    rng = ws.Range(ws.Cells(14, 2), ws.Cells(14, 5))
    align_cells(rng, (82, 170, 49),(255, 255, 255),  False)
    ws.Cells(15, 2).Value = 'Конечное сальдо 76.33 (ДДУ)'
    ws.Cells(15, 3).Value = f'=SUM({DDU_name}!AB:AB)'
    ws.Cells(15, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(15, 4).Value = f'=SUMIFS({BIT_row+len(df)+1}:{BIT_row+len(df)+1},{BIT_row-6}:{BIT_row-6},"ДДУ")'
    ws.Cells(15, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(15, 5).Value = f'=IF(ROUND(C15,1)=ROUND(D15,1),"КОРРЕКТНО","ОШИБКА")'
    ws.Cells(16, 2).Value = 'Оборот 90.01.1 (ДКП)'
    ws.Cells(16, 3).Value = f'=SUMIFS({DKP_name}!AB:AB,{DKP_name}!X:X,"Заключение")'
    ws.Cells(16, 3).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(16, 4).Value = f'=SUMIFS({BIT_row+len(df)+1}:{BIT_row+len(df)+1},{BIT_row-6}:{BIT_row-6},"ДКП")'
    ws.Cells(16, 4).NumberFormat = '# ###;(# ###);"-"'
    ws.Cells(16, 5).Value = f'=IF(ROUND(C16,1)=ROUND(D16,1),"КОРРЕКТНО","ОШИБКА")'
    rng = ws.Range(ws.Cells(15, 5), ws.Cells(16, 5))
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108

    set_column_width(ws, periods)
    rng = ws.Range(ws.Cells(8, 2), ws.Cells(12, 8))
    set_border(ws, rng, 8, 2, 12, 8, option=False)

    ws.Range(ws.Cells(12, 2), ws.Cells(12, 8)).Borders(8).Weight = 3
    ws.Range(ws.Cells(8, 2), ws.Cells(12, 2)).Borders(10).Weight = 2
    ws.Range(ws.Cells(8, 3), ws.Cells(12, 5)).Borders(10).Weight = 2
    set_conditional_formatting(ws.Range(ws.Cells(10, 6), ws.Cells(11, 8)))
    set_conditional_formatting(ws.Range(ws.Cells(15, 5), ws.Cells(16, 5)))
    rng = ws.Range(ws.Cells(14, 2), ws.Cells(16, 5))
    for border_id in range(7, 13):
        if border_id in (11, 12):
            rng.Borders(border_id).LineStyle = -4119
            rng.Borders(border_id).Weight = 1
        else:
            rng.Borders(border_id).LineStyle = 1
            rng.Borders(border_id).Weight = 2

def set_conditional_formatting(rng):
    positive_cond = rng.FormatConditions.Add(1, 3, "КОРРЕКТНО")
    positive_cond.Interior.Color = rgbToInt((198, 239, 206))
    positive_cond.Font.Color = rgbToInt((0, 97, 0))
    negative_cond = rng.FormatConditions.Add(1, 3, "ОШИБКА")
    negative_cond.Interior.Color = rgbToInt((255, 199, 206))
    negative_cond.Font.Color = rgbToInt((156, 0, 6))
def select_count_col(headname):
    if headname in ('Реализация по ДДУ', 'Разница'):
        return 5
    else:
        return 3
def rgbToInt(rgb):
    if isinstance(rgb, tuple):
        colorInt = rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)
        return colorInt
    else:
        return rgb


def decorate_cells(rng, interor_color, font_color, size, font_name, is_bold, is_underline =False, is_italic = False):
    rng.Interior.Color = rgbToInt(interor_color)
    rng.Font.Color = rgbToInt(font_color)
    rng.Font.Size = size
    rng.Font.Name = font_name
    rng.Font.Bold = is_bold
    rng.Font.Underline = is_underline
    rng.Font.Italic = is_italic
def align_cells(rng, interior_color, font_color, is_merge = True):
    rng.MergeCells = is_merge
    rng.VerticalAlignment = -4108
    rng.HorizontalAlignment = -4108
    decorate_cells(rng, rgbToInt(interior_color), rgbToInt(font_color), 10, 'Arial', True)

def set_column_width(ws, periods):
    for i in range(2, 10*len(periods)):
        if i == 2:
            ws.Columns(i).ColumnWidth = 29
        else:
            ws.Columns(i).ColumnWidth = 15

def set_border(ws, rng, init_row, init_col, last_row, last_col, color = '', option = True):
    for border_id in range(7, 13):
        if border_id in (11, 12):
            rng.Borders(border_id).LineStyle = -4119
            rng.Borders(border_id).Weight = 1
        else:
            rng.Borders(border_id).LineStyle = 1
            if border_id == 9:
                rng.Borders(border_id).Weight = 3
            else:
                rng.Borders(border_id).Weight = 2

    ws.Range(ws.Cells(init_row, init_col), ws.Cells(last_row, init_col)).Borders(8).Weight = 2
    if option:
        ws.Range((ws.Cells(last_row, init_col)), (ws.Cells(last_row, last_col))).Borders(8).Weight = 3
        ws.Range((ws.Cells(last_row-1, init_col)), (ws.Cells(last_row-1, last_col))).Borders(8).Weight = 2
        ws.Range((ws.Cells(last_row-1, init_col)), (ws.Cells(last_row-1, last_col))).Interior.Color = rgbToInt((112, 48, 160))
    else:
        ws.Range((ws.Cells(last_row, init_col)), (ws.Cells(last_row, last_col))).Borders(8).Weight = 3
        ws.Range(ws.Cells(init_row, init_col+1), ws.Cells(last_row, init_col+3)).Borders(10).Weight = 2

